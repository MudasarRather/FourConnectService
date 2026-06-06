"""TDS · Form 24Q Excel workbook — openpyxl renderer.

A bespoke quarterly TDS statement: an accent title rail, a dossier-purple KPI
tile strip, a PAN-wise body table that pairs *period* columns against *YTD*
columns, a traffic-light ColorScaleRule heat-map on the TDS-YTD column, an
effective-rate column, frozen header, autofilter, and a summing TOTAL row.

Engine: openpyxl (per report contract — period vs YTD layout + ColorScaleRule).
"""
from __future__ import annotations

import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter

from ..data import report_meta

MONEY_FMT = '"₹"#,##0'
PCT_FMT = '0.0"%"'

# dossier palette
ACCENT = "9333ea"
ACCENT_SOFT = "F3E8FF"
ACCENT_DEEP = "581C87"
INK = "1D1505"
CREAM = "FBF6EA"
DANGER = "B91C1C"
GOOD = "047857"
WHITE = "FFFFFF"


def _f(hexcol: str) -> str:
    return hexcol.lstrip("#").upper()


def render(rows: list[dict], summary: dict, meta: dict) -> bytes:
    theme = report_meta("tds-24q")
    accent = _f(theme.get("accent", ACCENT))
    deep = _f(theme.get("accent_deep", ACCENT_DEEP))
    period = meta.get("period", {})
    label = period.get("label", "")
    fy = period.get("fy", "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TDS - Form 24Q"[:31]
    ws.sheet_properties.tabColor = accent
    ws.sheet_view.showGridLines = False

    # ── Column model: PAN | Code | Name | (period) Taxable Gross, TDS |
    #                  (YTD) Gross, TDS | Eff. Rate %
    headers = [
        ("PAN", 16, "text"),
        ("Emp Code", 12, "text"),
        ("Deductee Name", 28, "text"),
        ("Taxable Gross\n(Period)", 16, "money"),
        ("TDS Deducted\n(Period)", 15, "money"),
        ("Gross\n(YTD)", 16, "money"),
        ("TDS\n(YTD)", 15, "money"),
        ("Eff. Rate\n(YTD %)", 11, "pct"),
    ]
    ncol = len(headers)
    last_letter = get_column_letter(ncol)
    for i, (_, w, _k) in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    thin = Side(style="thin", color="D8CDB5")
    rule_soft = Border(left=thin, right=thin, top=thin, bottom=thin)

    r = 1

    # ── Accent rail (height 4) ────────────────────────────────────────────
    ws.row_dimensions[r].height = 4
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor=accent)
    r += 1

    # ── Title row ─────────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 30
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    c = ws.cell(r, 1, f"Fourreck  ·  {theme.get('name', 'TDS · Form 24Q')}")
    c.font = Font(name="Calibri", size=18, bold=True, color=INK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.fill = PatternFill("solid", fgColor=WHITE)
    r += 1

    # ── Subtitle (italic) ─────────────────────────────────────────────────
    ws.row_dimensions[r].height = 18
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    c = ws.cell(r, 1, "  " + theme.get("subtitle", "PAN-wise tax deducted at source — period & year-to-date"))
    c.font = Font(name="Calibri", size=10, italic=True, color="475569")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = Border(bottom=Side(style="thin", color="CBD5E1"))
    r += 1

    # ── Period / FY row ───────────────────────────────────────────────────
    ws.row_dimensions[r].height = 20
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    c = ws.cell(r, 1, f"  Pay period   {label}      ·      FY {fy}      ·      Statutory Filing  ·  Quarterly 24Q")
    c.font = Font(name="Calibri", size=9, bold=True, color=INK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.fill = PatternFill("solid", fgColor="F1F5F9")
    c.border = Border(top=Side(style="thin", color="94A3B8"),
                      bottom=Side(style="medium", color=deep))
    r += 1

    # spacer
    ws.row_dimensions[r].height = 6
    r += 1

    # ── KPI tile strip ────────────────────────────────────────────────────
    employees = summary.get("employees", summary.get("rows", len(rows)))
    nrows = summary.get("rows", len(rows))
    tds_total = summary.get("tds", sum(float(x.get("tds_period") or 0) for x in rows))
    gross_total = summary.get("gross", sum(float(x.get("taxable_gross") or 0) for x in rows))

    kpis = [
        ("DEDUCTEES", employees, None, accent),
        ("PAN RECORDS", nrows, None, deep),
        ("TDS DEDUCTED", float(tds_total), MONEY_FMT, DANGER),
        ("TAXABLE GROSS", float(gross_total), MONEY_FMT, GOOD),
    ]
    # distribute KPI tiles across the columns (2 cols each over 8 cols)
    cols_per = max(1, ncol // len(kpis))
    label_row = r
    value_row = r + 1
    ws.row_dimensions[label_row].height = 18
    ws.row_dimensions[value_row].height = 28
    c0 = 1
    for i, (klab, kval, kfmt, krail) in enumerate(kpis):
        c1 = c0 + cols_per - 1
        if i == len(kpis) - 1:
            c1 = ncol
        ws.merge_cells(start_row=label_row, start_column=c0, end_row=label_row, end_column=c1)
        lc = ws.cell(label_row, c0, klab)
        lc.font = Font(name="Calibri", size=8, bold=True, color="475569")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.fill = PatternFill("solid", fgColor=WHITE)
        lc.border = Border(top=Side(style="thick", color=krail),
                           left=Side(style="thin", color="CBD5E1"),
                           right=Side(style="thin", color="CBD5E1"))

        ws.merge_cells(start_row=value_row, start_column=c0, end_row=value_row, end_column=c1)
        vc = ws.cell(value_row, c0, kval)
        vc.font = Font(name="Calibri", size=16, bold=True, color=INK)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.fill = PatternFill("solid", fgColor=WHITE)
        vc.border = Border(left=Side(style="thin", color="CBD5E1"),
                           right=Side(style="thin", color="CBD5E1"),
                           bottom=Side(style="thin", color="94A3B8"))
        if kfmt:
            vc.number_format = kfmt
        c0 = c1 + 1
    r = value_row + 1

    # spacer
    ws.row_dimensions[r].height = 8
    r += 1

    # ── Grouping band: PERIOD vs YEAR-TO-DATE ─────────────────────────────
    band_row = r
    ws.row_dimensions[band_row].height = 16
    # identity cols 1-3 blank band
    ws.merge_cells(start_row=band_row, start_column=1, end_row=band_row, end_column=3)
    bc = ws.cell(band_row, 1, "DEDUCTEE")
    bc.font = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
    bc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    bc.fill = PatternFill("solid", fgColor=deep)
    # period band cols 4-5
    ws.merge_cells(start_row=band_row, start_column=4, end_row=band_row, end_column=5)
    pc = ws.cell(band_row, 4, "THIS PERIOD")
    pc.font = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
    pc.alignment = Alignment(horizontal="center", vertical="center")
    pc.fill = PatternFill("solid", fgColor=accent)
    # YTD band cols 6-8
    ws.merge_cells(start_row=band_row, start_column=6, end_row=band_row, end_column=ncol)
    yc = ws.cell(band_row, 6, "YEAR-TO-DATE")
    yc.font = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
    yc.alignment = Alignment(horizontal="center", vertical="center")
    yc.fill = PatternFill("solid", fgColor=deep)
    r += 1

    # ── Header row ────────────────────────────────────────────────────────
    header_row = r
    ws.row_dimensions[header_row].height = 30
    head_fill = PatternFill("solid", fgColor=accent)
    head_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    head_border = Border(top=Side(style="medium", color=deep),
                         bottom=Side(style="medium", color=deep),
                         left=Side(style="thin", color=deep),
                         right=Side(style="thin", color=deep))
    for i, (htext, _w, kind) in enumerate(headers, start=1):
        cell = ws.cell(header_row, i, htext)
        cell.font = head_font
        cell.fill = head_fill
        align = "right" if kind in ("money", "pct") else "left"
        cell.alignment = Alignment(horizontal=align, vertical="center",
                                   wrap_text=True, indent=1 if align == "left" else 0)
        cell.border = head_border
    ws.freeze_panes = ws.cell(header_row + 1, 1)

    # ── Body rows ─────────────────────────────────────────────────────────
    first_data = header_row + 1
    rr = first_data
    for idx, row in enumerate(rows):
        zebra = idx % 2 == 1
        fill = PatternFill("solid", fgColor=CREAM) if zebra else PatternFill("solid", fgColor=WHITE)

        taxable = float(row.get("taxable_gross") or 0)
        tds_p = float(row.get("tds_period") or 0)
        gross_y = float(row.get("gross_ytd") or 0)
        tds_y = float(row.get("tds_ytd") or 0)
        eff = (tds_y / gross_y * 100.0) if gross_y else 0.0

        vals = [
            (row.get("pan") or "—", "text"),
            (row.get("employee_code") or "—", "text"),
            (row.get("employee_name") or "—", "text"),
            (taxable, "money"),
            (tds_p, "money"),
            (gross_y, "money"),
            (tds_y, "money"),
            (eff, "pct"),
        ]
        for i, (val, kind) in enumerate(vals, start=1):
            cell = ws.cell(rr, i, val)
            cell.fill = fill
            cell.border = rule_soft
            if kind == "text":
                mono = (i == 1)  # PAN in monospace-ish bold for officialdom
                cell.font = Font(name="Consolas" if mono else "Calibri", size=10,
                                 bold=mono, color=INK)
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            elif kind == "money":
                cell.font = Font(name="Calibri", size=10, color=INK)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = MONEY_FMT
            else:  # pct
                cell.font = Font(name="Calibri", size=10, bold=True, color=INK)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = PCT_FMT
        rr += 1
    last_data = rr - 1

    # ── TOTAL row ─────────────────────────────────────────────────────────
    if rows:
        tot_fill = PatternFill("solid", fgColor=deep)
        tot_font = Font(name="Calibri", size=10, bold=True, color="FDE68A")
        tot_border = Border(top=Side(style="medium", color=deep),
                            bottom=Side(style="medium", color=deep),
                            left=Side(style="thin", color=deep),
                            right=Side(style="thin", color=deep))
        ws.row_dimensions[rr].height = 20
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=3)
        tc = ws.cell(rr, 1, "TOTAL")
        tc.font = tot_font
        tc.fill = tot_fill
        tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        tc.border = tot_border
        for ci in (2, 3):
            cc = ws.cell(rr, ci)
            cc.fill = tot_fill
            cc.border = tot_border
        # money sums cols 4..7
        for ci in (4, 5, 6, 7):
            col = get_column_letter(ci)
            cell = ws.cell(rr, ci, f"=SUM({col}{first_data}:{col}{last_data})")
            cell.font = tot_font
            cell.fill = tot_fill
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = MONEY_FMT
            cell.border = tot_border
        # blended effective rate col 8 = TDS_ytd / Gross_ytd
        gcol, tcol = get_column_letter(6), get_column_letter(7)
        eff_cell = ws.cell(rr, 8,
                           f'=IF(SUM({gcol}{first_data}:{gcol}{last_data})=0,0,'
                           f'SUM({tcol}{first_data}:{tcol}{last_data})/'
                           f'SUM({gcol}{first_data}:{gcol}{last_data})*100)')
        eff_cell.font = tot_font
        eff_cell.fill = tot_fill
        eff_cell.alignment = Alignment(horizontal="right", vertical="center")
        eff_cell.number_format = PCT_FMT
        eff_cell.border = tot_border

    # ── Conditional formats ───────────────────────────────────────────────
    if rows:
        # traffic-light heat-map on TDS (YTD) — col 7
        rng_tds_ytd = f"G{first_data}:G{last_data}"
        ws.conditional_formatting.add(
            rng_tds_ytd,
            ColorScaleRule(
                start_type="min", start_color="E6FFEC",
                mid_type="percentile", mid_value=50, mid_color="FEF9C3",
                end_type="max", end_color="FCA5A5",
            ),
        )
        # high effective-rate flag on col 8 (>= 20%)
        rng_eff = f"H{first_data}:H{last_data}"
        ws.conditional_formatting.add(
            rng_eff,
            CellIsRule(operator="greaterThanOrEqual", formula=["20"],
                       fill=PatternFill("solid", fgColor="FEE2E2"),
                       font=Font(color="7F1D1D", bold=True)),
        )

        ws.auto_filter.ref = f"A{header_row}:{last_letter}{last_data}"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()

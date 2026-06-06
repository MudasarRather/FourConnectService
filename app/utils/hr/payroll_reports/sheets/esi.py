"""ESI Contribution Statement workbook — report key "esi".

openpyxl renderer with an "insurance-card" feel: a blue ESIC accent rail, a
title block + period line, KPI tiles with thick coloured top borders, a frozen
themed header, bordered zebra body rows, monospace IP / ESIC numbers, a money
total row, and a ColorScaleRule traffic-light on the per-member total
contribution column. Matches the PDF's #0369a1 officialdom accent.

Public entry point::

    render(rows, summary, meta) -> bytes
"""
from __future__ import annotations

import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

from ..data import report_meta


MONEY_FMT = '"₹"#,##0'
DAYS_FMT = '0.0'

# Column layout (1-based): ESIC No · Code · Member · ESI Wages · EE 0.75% ·
# ER 3.25% · Total · Paid Days
COLS = [
    ("ESIC IP Number", "esic_number", 18, "mono"),
    ("Emp Code", "employee_code", 12, "mono"),
    ("Insured Member", "employee_name", 30, "text"),
    ("ESI Wages", "esi_wages", 14, "money"),
    ("EE @ 0.75%", "ee_esi", 13, "money"),
    ("ER @ 3.25%", "er_esi", 13, "money"),
    ("Total Contribution", "total_esi", 17, "money"),
    ("Paid Days", "paid_days", 11, "days"),
]
N = len(COLS)


def _f(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def render(rows: list[dict], summary: dict, meta: dict) -> bytes:
    theme = report_meta("esi")
    accent = theme.get("accent", "#0369a1")
    deep = theme.get("accent_deep", "#0c4a6e")
    soft = "e0f2fe"
    name = theme.get("name", "ESI Contribution Statement")
    subtitle = theme.get("subtitle", "Insurable wages · 0.75% EE · 3.25% ER per member")
    period = (meta or {}).get("period", {}) or {}

    A = accent.lstrip("#")
    D = deep.lstrip("#")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name[:31]
    ws.sheet_properties.tabColor = A
    ws.sheet_view.showGridLines = False

    # column widths
    for i, (_, _, w, _) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    thin = Side(style="thin", color="BFD9EC")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    last_letter = get_column_letter(N)

    # ── Row 1: accent rail ──
    ws.row_dimensions[1].height = 4
    rail_fill = PatternFill("solid", fgColor=A)
    for c in range(1, N + 1):
        ws.cell(row=1, column=c).fill = rail_fill
    ws.merge_cells(f"A1:{last_letter}1")

    # ── Row 2: government authority line ──
    ws.row_dimensions[2].height = 16
    ws.merge_cells(f"A2:{last_letter}2")
    gov = ws.cell(row=2, column=1,
                  value="EMPLOYEES' STATE INSURANCE CORPORATION · Govt. of India · Monthly Contribution Return")
    gov.font = Font(name="Calibri", size=8, bold=True, color=D)
    gov.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    gov.fill = PatternFill("solid", fgColor="FFFFFF")

    # ── Row 3: title ──
    ws.row_dimensions[3].height = 30
    ws.merge_cells(f"A3:{last_letter}3")
    t = ws.cell(row=3, column=1, value=f"  {name}")
    t.font = Font(name="Calibri", size=18, bold=True, color="0B2942")
    t.alignment = Alignment(horizontal="left", vertical="center")
    t.fill = PatternFill("solid", fgColor="FFFFFF")

    # ── Row 4: subtitle ──
    ws.row_dimensions[4].height = 18
    ws.merge_cells(f"A4:{last_letter}4")
    st = ws.cell(row=4, column=1, value=f"  {subtitle}")
    st.font = Font(name="Calibri", size=10, italic=True, color=D)
    st.alignment = Alignment(horizontal="left", vertical="center")
    st.fill = PatternFill("solid", fgColor="FFFFFF")

    # ── Row 5: period / employer line ──
    ws.row_dimensions[5].height = 20
    ws.merge_cells(f"A5:{last_letter}5")
    plabel = period.get("label", "")
    fy = period.get("fy", "")
    pl = ws.cell(
        row=5, column=1,
        value=(f"  Employer Code  31-00098765-001      ·      Contribution Month  {plabel}"
               f"      ·      FY {fy}      ·      Wage Ceiling  ₹21,000"),
    )
    pl.font = Font(name="Calibri", size=9, bold=True, color="0B2942")
    pl.alignment = Alignment(horizontal="left", vertical="center")
    pl.fill = PatternFill("solid", fgColor=soft)
    pl.border = Border(bottom=Side(style="thick", color=D))

    # ── KPI tiles (rows 7 label / 8 value) ──
    esi_total = _f(summary.get("esi"))
    ee_share = esi_total * (0.75 / 4.0)
    er_share = esi_total * (3.25 / 4.0)
    members = summary.get("rows", summary.get("employees", len(rows)))
    employees = summary.get("employees", members)

    kpis = [
        ("INSURED MEMBERS", members, A, None),
        ("EMPLOYEES", employees, D, None),
        ("EE SHARE (0.75%)", ee_share, "0E7490", MONEY_FMT),
        ("ER SHARE (3.25%)", er_share, "0F766E", MONEY_FMT),
        ("TOTAL PAYABLE", esi_total, A, MONEY_FMT),
    ]
    klabel_row, kval_row = 7, 8
    ws.row_dimensions[6].height = 6
    ws.row_dimensions[klabel_row].height = 16
    ws.row_dimensions[kval_row].height = 26

    # distribute KPI tiles across the N columns
    per = max(1, N // len(kpis))
    col = 1
    for idx, (klab, kval, rail, numfmt) in enumerate(kpis):
        span = per if idx < len(kpis) - 1 else (N - (col - 1))
        c0, c1 = col, min(col + span - 1, N)
        l0 = get_column_letter(c0)
        l1 = get_column_letter(c1)
        ws.merge_cells(f"{l0}{klabel_row}:{l1}{klabel_row}")
        ws.merge_cells(f"{l0}{kval_row}:{l1}{kval_row}")

        lc = ws.cell(row=klabel_row, column=c0, value=klab)
        lc.font = Font(name="Calibri", size=8, bold=True, color="475569")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.fill = PatternFill("solid", fgColor="FFFFFF")
        lc.border = Border(top=Side(style="thick", color=rail),
                           left=Side(style="thin", color="CBD5E1"),
                           right=Side(style="thin", color="CBD5E1"))

        vc = ws.cell(row=kval_row, column=c0, value=kval)
        vc.font = Font(name="Calibri", size=15, bold=True, color="0B2942")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.fill = PatternFill("solid", fgColor="FFFFFF")
        vc.border = Border(left=Side(style="thin", color="CBD5E1"),
                           right=Side(style="thin", color="CBD5E1"),
                           bottom=Side(style="thin", color="CBD5E1"))
        if numfmt:
            vc.number_format = numfmt
        col = c1 + 1

    ws.row_dimensions[9].height = 6

    # ── Header row ──
    header_row = 10
    ws.row_dimensions[header_row].height = 26
    head_fill = PatternFill("solid", fgColor=A)
    head_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    head_border = Border(left=Side(style="thin", color=D), right=Side(style="thin", color=D),
                         top=Side(style="medium", color=D), bottom=Side(style="medium", color=D))
    for i, (label, _, _, kind) in enumerate(COLS, start=1):
        hc = ws.cell(row=header_row, column=i, value=label)
        hc.font = head_font
        hc.fill = head_fill
        hc.alignment = Alignment(
            horizontal="right" if kind in ("money", "days") else "left",
            vertical="center", indent=1, wrap_text=False)
        hc.border = head_border

    # ── Body rows ──
    first_data = header_row + 1
    mono_font = Font(name="Consolas", size=10, color="0B2942")
    text_font = Font(name="Calibri", size=10, color="111418")
    num_font = Font(name="Calibri", size=10, color="111418")
    zebra_fill = PatternFill("solid", fgColor="F4FAFE")

    for ri, r in enumerate(rows):
        rownum = first_data + ri
        ws.row_dimensions[rownum].height = 17
        zebra = ri % 2 == 1
        for i, (_, key, _, kind) in enumerate(COLS, start=1):
            cell = ws.cell(row=rownum, column=i)
            if kind == "money":
                cell.value = _f(r.get(key))
                cell.number_format = MONEY_FMT
                cell.font = num_font
                cell.alignment = Alignment(horizontal="right", indent=1, vertical="center")
            elif kind == "days":
                cell.value = _f(r.get(key))
                cell.number_format = DAYS_FMT
                cell.font = num_font
                cell.alignment = Alignment(horizontal="right", indent=1, vertical="center")
            elif kind == "mono":
                v = r.get(key)
                cell.value = v if v not in (None, "") else "—"
                cell.font = mono_font
                cell.alignment = Alignment(horizontal="left", indent=1, vertical="center")
            else:
                v = r.get(key)
                cell.value = v if v not in (None, "") else "—"
                cell.font = text_font
                cell.alignment = Alignment(horizontal="left", indent=1, vertical="center")
            cell.border = border_all
            if zebra:
                cell.fill = zebra_fill

    last_data = first_data + len(rows) - 1 if rows else header_row

    # ── Total row (sum money columns) ──
    total_row = (last_data + 1) if rows else (header_row + 1)
    ws.row_dimensions[total_row].height = 20
    tot_fill = PatternFill("solid", fgColor=D)
    tot_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    for i, (label, key, _, kind) in enumerate(COLS, start=1):
        cell = ws.cell(row=total_row, column=i)
        cell.fill = tot_fill
        cell.font = tot_font
        cell.border = Border(top=Side(style="medium", color=D), bottom=Side(style="medium", color=D))
        if i == 1:
            cell.value = "TOTAL"
            cell.alignment = Alignment(horizontal="left", indent=1, vertical="center")
        elif kind == "money" and rows:
            col_letter = get_column_letter(i)
            cell.value = f"=SUM({col_letter}{first_data}:{col_letter}{last_data})"
            cell.number_format = MONEY_FMT
            cell.alignment = Alignment(horizontal="right", indent=1, vertical="center")
        elif key == "paid_days" and rows:
            col_letter = get_column_letter(i)
            cell.value = f"=SUM({col_letter}{first_data}:{col_letter}{last_data})"
            cell.number_format = DAYS_FMT
            cell.alignment = Alignment(horizontal="right", indent=1, vertical="center")

    # ── Freeze header + ESIC/Code/Member key columns ──
    ws.freeze_panes = ws.cell(row=first_data, column=4)

    # ── Autofilter ──
    if rows:
        ws.auto_filter.ref = f"A{header_row}:{last_letter}{last_data}"

        # Traffic-light ColorScaleRule on the Total Contribution column (col 7)
        total_col = get_column_letter(7)
        rng = f"{total_col}{first_data}:{total_col}{last_data}"
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="min", start_color="E0F2FE",
                mid_type="percentile", mid_value=50, mid_color="7DD3FC",
                end_type="max", end_color="0369A1",
            ),
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

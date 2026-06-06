"""Excel renderer for the Year-to-Date Earnings report (openpyxl).

A year-end statement: accent title rail, ceremonial KPI tiles with colored top
borders, a themed frozen header, bordered zebra body, a money-formatted TOTAL
row, and a traffic-light ColorScaleRule heat-map on the YTD Net column. Engine
is openpyxl (edit-friendly, formula-aware) per the report brief.
"""
from __future__ import annotations

import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter

from ..data import report_meta


MONEY_FMT = '"₹"#,##0'
PCT_FMT = '0.0"%"'

SHEET_NAME = "Year-to-Date Earnings"


# (header, row-key, width, kind)  kind ∈ text|num|money|pct|months
_COLUMNS = [
    ("Emp Code",        "employee_code",   13, "text"),
    ("Employee Name",   "employee_name",   28, "text"),
    ("Department",      "department",       20, "text"),
    ("Months Paid",     "months_paid",      12, "months"),
    ("YTD Gross",       "ytd_gross",        15, "money"),
    ("YTD Deductions",  "ytd_deductions",   16, "money"),
    ("YTD TDS",         "ytd_tds",          14, "money"),
    ("YTD PF",          "ytd_pf",           13, "money"),
    ("YTD Net",         "ytd_net",          15, "money"),
    ("Employer Cost",   "ytd_employer",     15, "money"),
    ("Take-home %",     None,               13, "pct"),   # computed net/gross
]


def render(rows: list[dict], summary: dict, meta: dict) -> bytes:
    theme = report_meta("ytd-earnings")
    accent = theme["accent"]          # #92400e
    deep = theme["accent_deep"]       # #451a03
    name = theme.get("name", "Year-to-Date Earnings")
    subtitle = theme.get("subtitle", "")
    period = meta.get("period", {}) or {}

    gold = "b8860b"
    gold_lt = "d9a441"
    cream = "fbf8f0"
    cream2 = "fff9ec"
    a = accent.lstrip("#")
    d = deep.lstrip("#")

    ncols = len(_COLUMNS)
    last_letter = get_column_letter(ncols)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME[:31]
    ws.sheet_properties.tabColor = a
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin", color="d8cdb5")
    thin_g = Side(style="thin", color=gold)

    # column widths
    for i, (_, _, w, _) in enumerate(_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    # ── Accent rail (height 4) ──
    rail_fill = PatternFill("solid", fgColor=a)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    for c in range(1, ncols + 1):
        ws.cell(row=r, column=c).fill = rail_fill
    ws.row_dimensions[r].height = 4
    r += 1

    # ── Title row (18pt bold) ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    tcell = ws.cell(row=r, column=1, value=f"Fourreck  ·  {name}")
    tcell.font = Font(name="Georgia", size=18, bold=True, color=d)
    tcell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    tcell.fill = PatternFill("solid", fgColor="ffffff")
    ws.row_dimensions[r].height = 30
    r += 1

    # ── Subtitle (italic) ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    scell = ws.cell(row=r, column=1, value=subtitle)
    scell.font = Font(name="Georgia", size=10, italic=True, color="5a4632")
    scell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    scell.border = Border(bottom=Side(style="thin", color=gold_lt))
    ws.row_dimensions[r].height = 18
    r += 1

    # ── Period row ──
    from datetime import datetime
    gen = datetime.now().strftime("%d %b %Y · %I:%M %p").lstrip("0")
    pline = (f"Fiscal Year  {period.get('fy', '')}     ·     Cumulative through  "
             f"{period.get('label', '')}     ·     Generated  {gen}")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    pcell = ws.cell(row=r, column=1, value=pline)
    pcell.font = Font(name="Calibri", size=9, bold=True, color=d)
    pcell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    pcell.fill = PatternFill("solid", fgColor="f3e2b3")
    pcell.border = Border(bottom=Side(style="medium", color=d))
    ws.row_dimensions[r].height = 20
    r += 2

    # ── KPI tiles (each spans ~2-3 cols, colored top border via thick Side) ──
    kpis = [
        ("EMPLOYEES",      summary.get("employees", len(rows)), d,       None),
        ("YTD GROSS",      summary.get("ytd_gross", 0),         gold,    MONEY_FMT),
        ("YTD NET",        summary.get("ytd_net", 0),           "047857", MONEY_FMT),
        ("YTD TDS",        summary.get("ytd_tds", 0),           "b91c1c", MONEY_FMT),
    ]
    nk = len(kpis)
    span = max(1, ncols // nk)
    label_row = r
    value_row = r + 1
    ws.row_dimensions[label_row].height = 18
    ws.row_dimensions[value_row].height = 30
    col = 1
    for i, (lbl, val, color, fmt) in enumerate(kpis):
        c1 = col
        c2 = (ncols if i == nk - 1 else min(col + span - 1, ncols))
        # label
        ws.merge_cells(start_row=label_row, start_column=c1, end_row=label_row, end_column=c2)
        lc = ws.cell(row=label_row, column=c1, value=lbl)
        lc.font = Font(name="Calibri", size=8, bold=True, color="475569")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.fill = PatternFill("solid", fgColor=cream2)
        lc.border = Border(top=Side(style="thick", color=color),
                           left=Side(style="thin", color=gold),
                           right=Side(style="thin", color=gold))
        # value
        ws.merge_cells(start_row=value_row, start_column=c1, end_row=value_row, end_column=c2)
        vc = ws.cell(row=value_row, column=c1, value=val)
        vc.font = Font(name="Calibri", size=16, bold=True, color=d)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.fill = PatternFill("solid", fgColor=cream2)
        vc.border = Border(left=Side(style="thin", color=gold),
                           right=Side(style="thin", color=gold),
                           bottom=Side(style="thin", color=gold))
        if fmt:
            vc.number_format = fmt
        col = c2 + 1
    r = value_row + 2

    # ── Header row ──
    head_fill = PatternFill("solid", fgColor=a)
    head_border = Border(top=Side(style="medium", color=d), bottom=Side(style="medium", color=d),
                         left=Side(style="thin", color=d), right=Side(style="thin", color=d))
    header_row = r
    for i, (label, _, _, kind) in enumerate(_COLUMNS, start=1):
        hc = ws.cell(row=header_row, column=i, value=label)
        hc.font = Font(name="Calibri", size=10, bold=True, color="ffffff")
        hc.fill = head_fill
        hc.border = head_border
        hc.alignment = Alignment(
            horizontal=("right" if kind in ("money", "num", "pct", "months") else "left"),
            vertical="center", indent=1, wrap_text=False)
    ws.row_dimensions[header_row].height = 26
    ws.freeze_panes = ws.cell(row=header_row + 1, column=4)  # freeze header + first 3 id cols

    # ── Body rows ──
    body_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    first_data = header_row + 1
    for ri, row in enumerate(rows):
        rr = first_data + ri
        zebra = ri % 2 == 1
        fill = PatternFill("solid", fgColor=cream) if zebra else PatternFill("solid", fgColor="ffffff")
        gross = float(row.get("ytd_gross") or 0)
        net = float(row.get("ytd_net") or 0)
        take_home = (net / gross * 100.0) if gross else 0.0
        for i, (label, key, _, kind) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=rr, column=i)
            cell.fill = fill
            cell.border = body_border
            if kind == "pct":
                cell.value = round(take_home, 1)
                cell.number_format = PCT_FMT
                cell.alignment = Alignment(horizontal="right", indent=1, vertical="center")
                cell.font = Font(name="Calibri", size=10, bold=True, color="047857")
            elif kind == "money":
                cell.value = float(row.get(key) or 0)
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right", indent=1, vertical="center")
                cell.font = Font(name="Calibri", size=10, color="1a1410")
            elif kind == "months":
                cell.value = int(row.get(key) or 0)
                cell.number_format = '0" mo"'
                cell.alignment = Alignment(horizontal="right", indent=1, vertical="center")
                cell.font = Font(name="Calibri", size=10, color="1a1410")
            else:  # text
                v = row.get(key)
                cell.value = v if v not in (None, "") else "—"
                cell.alignment = Alignment(horizontal="left", indent=1, vertical="center")
                bold = (key == "employee_name")
                cell.font = Font(name="Calibri", size=10, bold=bold, color="1a1410")

    last_data = first_data + len(rows) - 1 if rows else header_row

    # ── TOTAL row ──
    if rows:
        tr = last_data + 1
        total_fill = PatternFill("solid", fgColor=d)
        total_border = Border(top=Side(style="medium", color=d), bottom=Side(style="medium", color=d),
                              left=Side(style="thin", color=d), right=Side(style="thin", color=d))
        ws.row_dimensions[tr].height = 22
        money_keys = {"ytd_gross", "ytd_deductions", "ytd_tds", "ytd_pf", "ytd_net", "ytd_employer"}
        for i, (label, key, _, kind) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=tr, column=i)
            cell.fill = total_fill
            cell.border = total_border
            colL = get_column_letter(i)
            if i == 1:
                cell.value = "TOTAL"
                cell.font = Font(name="Calibri", size=10, bold=True, color="ffffff")
                cell.alignment = Alignment(horizontal="left", indent=1, vertical="center")
            elif key in money_keys:
                cell.value = f"=SUM({colL}{first_data}:{colL}{last_data})"
                cell.number_format = MONEY_FMT
                cell.font = Font(name="Calibri", size=10, bold=True, color="ffffff")
                cell.alignment = Alignment(horizontal="right", indent=1, vertical="center")
            elif kind == "pct":
                gcol = get_column_letter(_money_col("ytd_gross"))
                ncol = get_column_letter(_money_col("ytd_net"))
                cell.value = f"=IF({gcol}{tr}=0,0,{ncol}{tr}/{gcol}{tr}*100)"
                cell.number_format = PCT_FMT
                cell.font = Font(name="Calibri", size=10, bold=True, color="ffffff")
                cell.alignment = Alignment(horizontal="right", indent=1, vertical="center")
            elif kind == "months":
                cell.value = f"=MAX({colL}{first_data}:{colL}{last_data})"
                cell.number_format = '0" mo"'
                cell.font = Font(name="Calibri", size=10, bold=True, color="ffffff")
                cell.alignment = Alignment(horizontal="right", indent=1, vertical="center")
            else:
                cell.value = None

    # ── ColorScaleRule heat-map on YTD Net column ──
    if rows:
        net_col = get_column_letter(_money_col("ytd_net"))
        rng = f"{net_col}{first_data}:{net_col}{last_data}"
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="min", start_color="fde8cf",
            mid_type="percentile", mid_value=50, mid_color="f0b860",
            end_type="max", end_color=gold,
        ))
        # Take-home % traffic light: flag <70% as a concern
        pct_col = get_column_letter(len(_COLUMNS))
        prng = f"{pct_col}{first_data}:{pct_col}{last_data}"
        ws.conditional_formatting.add(prng, CellIsRule(
            operator="lessThan", formula=["70"],
            fill=PatternFill("solid", fgColor="fee2e2"),
            font=Font(color="7f1d1d", bold=True)))
        ws.conditional_formatting.add(prng, CellIsRule(
            operator="greaterThanOrEqual", formula=["85"],
            fill=PatternFill("solid", fgColor="dcfce7"),
            font=Font(color="166534", bold=True)))

    # ── Autofilter over header + body ──
    if rows:
        ws.auto_filter.ref = f"A{header_row}:{last_letter}{last_data}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _money_col(key: str) -> int:
    for i, (_, k, _, _) in enumerate(_COLUMNS, start=1):
        if k == key:
            return i
    return 1

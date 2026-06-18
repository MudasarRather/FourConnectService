"""HR Training Reports — Excel exporter (openpyxl).

Two-sheet workbook, deliberately "ultra-modern with proper data":

    • Overview — a branded KPI board (big-number tiles from the report summary)
      with the report's accent, period strip and generated stamp.
    • Data — a styled, sortable table: dark accent header, zebra rows, real
      typed values (numbers/dates, not strings) so totals/sorting work,
      money/percent number-formats, traffic-light conditional formatting on
      rate/gap columns, frozen header, autofilter, and a native bar chart for
      the report's headline metric.

openpyxl (not xlsxwriter) so we get conditional formatting + can keep charts in
the same workbook; the chart is a simple BarChart guarded in try/except.
"""
from __future__ import annotations

import io
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.chart import BarChart, Reference

from .data import report_meta

_INK = "1A1410"
_MUTE = "6B5840"
_GRID = "ECE3D2"
_PAPER = "FFFDF9"     # warm near-white sheet base
_HEAD_TXT = "FFF3D6"  # warm cream for text on the dark header band


def _hx(c: str) -> str:
    return c.lstrip("#").upper()


def _outline(ws, r1, c1, r2, c2, side):
    """Draw a border on the *perimeter* of a rectangular range, preserving any
    interior borders already set on the cells."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            b = cell.border
            cell.border = Border(
                top=side if r == r1 else b.top,
                bottom=side if r == r2 else b.bottom,
                left=side if c == c1 else b.left,
                right=side if c == c2 else b.right,
            )


def _grid(ws, r1, c1, r2, c2, side):
    """Set a thin border on all four edges of every cell in the range."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = Border(top=side, bottom=side, left=side, right=side)


def _fill_rect(ws, r1, c1, r2, c2, fill):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).fill = fill


def _draw_tile(ws, top, lcol, label, value, fmt, accent, deep, soft):
    """Render one KPI card spanning two columns × four rows:
    accent strip · label · big value · pad — bordered like a dashboard tile."""
    rcol = lcol + 1
    soft_fill = PatternFill("solid", fgColor=soft)
    accent_fill = PatternFill("solid", fgColor=accent)

    # body fill + accent strip
    _fill_rect(ws, top, lcol, top + 3, rcol, soft_fill)
    _fill_rect(ws, top, lcol, top, rcol, accent_fill)

    # merge each band across the two columns
    for rr in range(top, top + 4):
        ws.merge_cells(start_row=rr, start_column=lcol, end_row=rr, end_column=rcol)

    lab = ws.cell(row=top + 1, column=lcol, value=str(label).upper())
    lab.font = Font(bold=True, size=8, color=_MUTE)
    lab.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    val = ws.cell(row=top + 2, column=lcol, value=_coerce(value, fmt))
    val.font = Font(bold=True, size=22, color=deep)
    val.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if fmt in _NUMFMT and isinstance(value, (int, float)):
        val.number_format = _NUMFMT[fmt]

    # row heights + soft card border
    ws.row_dimensions[top].height = 5
    ws.row_dimensions[top + 1].height = 16
    ws.row_dimensions[top + 2].height = 30
    ws.row_dimensions[top + 3].height = 7
    _outline(ws, top, lcol, top + 3, rcol, Side(style="thin", color=_GRID))


def _summary_tiles(report: dict) -> list:
    """Pick the headline KPIs from each report's summary as (label, value, fmt)."""
    s = report.get("summary", {})
    k = report["key"]
    tiles = {
        "enrollments": [("Enrollments", s.get("total"), "int"), ("Completed", s.get("completed"), "int"),
                        ("In progress", s.get("in_progress"), "int"), ("Overdue", s.get("overdue"), "int"),
                        ("Completion", s.get("completion_rate"), "pct")],
        "completion": [("Programs", s.get("programs"), "int"), ("Enrolled", s.get("enrolled"), "int"),
                       ("Completed", s.get("completed"), "int"), ("Completion", s.get("completion_rate"), "pct")],
        "assessments": [("Assessments", s.get("assessments"), "int"), ("Attempts", s.get("attempts"), "int"),
                        ("Passed", s.get("passed"), "int"), ("Pass rate", s.get("pass_rate"), "pct"),
                        ("Avg score", s.get("avg_score"), "pct")],
        "feedback": [("Programs", s.get("programs"), "int"), ("Responses", s.get("responses"), "int"),
                     ("Avg rating", s.get("avg_rating"), "num")],
        "skill_gap": [("Skills", s.get("skills"), "int"), ("Avg gap", s.get("avg_gap"), "num"),
                      ("With gap", s.get("with_gap"), "int"), ("Critical", s.get("critical"), "int")],
        "certifications": [("Credentials", s.get("total"), "int"), ("Active", s.get("active"), "int"),
                           ("Expiring", s.get("expiring"), "int"), ("Expired", s.get("expired"), "int")],
        "trainers": [("Trainers", s.get("trainers"), "int"), ("Active", s.get("active"), "int"),
                     ("Avg rating", s.get("avg_rating"), "num"), ("Ratings", s.get("responses"), "int")],
        "compliance": [("Programs", s.get("programs"), "int"), ("Eligible", s.get("eligible"), "int"),
                       ("Compliant", s.get("compliant"), "int"), ("Overdue", s.get("overdue"), "int"),
                       ("Coverage", s.get("coverage"), "pct")],
        "requests": [("Requests", s.get("total"), "int"), ("Pending", s.get("pending"), "int"),
                     ("Fulfilled", s.get("fulfilled"), "int"), ("Rejected", s.get("rejected"), "int"),
                     ("Fulfil rate", s.get("fulfil_rate"), "pct")],
        "budget": [("Budgets", s.get("budgets"), "int"), ("Allocated", s.get("allocated"), "money"),
                   ("Spent", s.get("spent"), "money"), ("Committed", s.get("committed"), "money"),
                   ("Utilization", s.get("utilization"), "pct")],
        "department": [("Departments", s.get("departments"), "int"), ("People", s.get("employees"), "int"),
                       ("Assigned", s.get("assignments"), "int"), ("Completion", s.get("completion_rate"), "pct"),
                       ("Active certs", s.get("active_certs"), "int")],
        # ── self-service ──
        "my_record": [("Programs", s.get("total"), "int"), ("Completed", s.get("completed"), "int"),
                      ("In progress", s.get("in_progress"), "int"), ("Overdue", s.get("overdue"), "int"),
                      ("Completion", s.get("completion_rate"), "pct"), ("Learning hours", s.get("hours"), "num")],
        "my_skills": [("Skills", s.get("skills"), "int"), ("At target", s.get("at_target"), "int"),
                      ("With gap", s.get("with_gap"), "int"), ("Mastered", s.get("mastered"), "int"),
                      ("Avg gap", s.get("avg_gap"), "num")],
        "my_credentials": [("Credentials", s.get("total"), "int"), ("Active", s.get("active"), "int"),
                           ("Expiring", s.get("expiring"), "int"), ("Expired", s.get("expired"), "int")],
        "my_requests": [("Requests", s.get("total"), "int"), ("Pending", s.get("pending"), "int"),
                        ("Approved", s.get("approved"), "int"), ("Fulfilled", s.get("fulfilled"), "int"),
                        ("Fulfil rate", s.get("fulfil_rate"), "pct")],
    }
    return tiles.get(k, [(kk.replace("_", " ").title(), vv, "int") for kk, vv in list(s.items())[:5]])


def _coerce(value, fmt):
    """Return a real Excel-typed value where possible (numbers/dates)."""
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if fmt in ("money", "pct", "num", "int") and isinstance(value, (int, float)):
        return value
    return value


_NUMFMT = {"money": '₹#,##0', "pct": '0.0"%"', "num": '0.00', "int": '#,##0'}


def render_excel(report: dict) -> bytes:
    meta = report_meta(report["key"])
    accent = _hx(meta["accent"]); deep = _hx(meta["accent_deep"]); soft = _hx(meta["accent_soft"])
    cols = report["columns"]
    rows = report["rows"]

    wb = Workbook()
    paper = PatternFill("solid", fgColor=_PAPER)

    # ───────────────────────── Overview sheet ─────────────────────────
    ov = wb.active
    ov.title = "Overview"
    ov.sheet_view.showGridLines = False
    ov.sheet_properties.tabColor = accent
    # layout: gutter A, then 4 tiles of 2 cols each separated by a thin gap col
    ov.column_dimensions["A"].width = 2.4
    tile_cols = []  # left column of each of the 4 tile slots
    for slot in range(4):
        lc = 2 + slot * 3
        tile_cols.append(lc)
        ov.column_dimensions[get_column_letter(lc)].width = 13
        ov.column_dimensions[get_column_letter(lc + 1)].width = 13
        ov.column_dimensions[get_column_letter(lc + 2)].width = 2.6
    last_grid_col = tile_cols[-1] + 1  # col L
    # warm paper wash behind the dashboard
    _fill_rect(ov, 1, 1, 60, last_grid_col + 1, paper)

    band = ov.cell(row=2, column=2, value=meta["eyebrow"])
    band.font = Font(bold=True, size=9, color=deep)
    t = ov.cell(row=3, column=2, value=report["title"])
    t.font = Font(bold=True, size=24, color=_INK)
    ov.row_dimensions[3].height = 30
    st = ov.cell(row=4, column=2, value=report.get("subtitle", ""))
    st.font = Font(italic=True, size=11, color=_MUTE)
    pr = ov.cell(row=5, column=2,
                 value=f"Period: {report.get('period', {}).get('label', 'All time')}   ·   "
                       f"Generated { datetime.now().strftime('%d %b %Y · %I:%M %p').lstrip('0') }")
    pr.font = Font(size=9, color=_MUTE)

    # thick accent rule under the header
    _fill_rect(ov, 6, 2, 6, last_grid_col, PatternFill("solid", fgColor=accent))
    ov.row_dimensions[6].height = 5

    # KPI cards (4 per row, two columns each)
    tiles = [tl for tl in _summary_tiles(report) if tl[1] is not None]
    r0 = 8
    for i, (label, value, fmt) in enumerate(tiles):
        lcol = tile_cols[i % 4]
        top = r0 + (i // 4) * 6  # 4 tile rows + a 2-row gap
        _draw_tile(ov, top, lcol, label, value, fmt, accent, deep, soft)

    foot_row = r0 + ((len(tiles) - 1) // 4) * 6 + 6
    foot = ov.cell(row=foot_row, column=2,
                   value="FourConnect HRMS — Training & Development · Confidential, internal use only")
    foot.font = Font(size=8, color="9A8A72")

    # ───────────────────────── Data sheet ─────────────────────────
    ws = wb.create_sheet("Data")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = deep
    ncols = len(cols)
    last_col_letter = get_column_letter(ncols)
    header_fill = PatternFill("solid", fgColor=_INK)
    header_font = Font(bold=True, color=_HEAD_TXT, size=10)
    thin = Side(style="thin", color=_GRID)
    med = Side(style="medium", color=deep)
    zebra = PatternFill("solid", fgColor=soft)
    white = PatternFill("solid", fgColor=_PAPER)

    # title band on data sheet
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    dt = ws.cell(row=1, column=1, value=report["title"])
    dt.font = Font(bold=True, size=14, color=deep)
    dt.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    _fill_rect(ws, 1, 1, 1, ncols, PatternFill("solid", fgColor=soft))
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    sub = ws.cell(row=2, column=1,
                  value=f"{report.get('period', {}).get('label', 'All time')}  ·  {len(rows)} record(s)")
    sub.font = Font(size=9, color=_MUTE, italic=True)
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    hdr_row = 3
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=hdr_row, column=i, value=c["label"])
        cell.fill = header_fill; cell.font = header_font
        align = "right" if c.get("align") == "right" else "center" if c.get("align") == "center" else "left"
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    ws.row_dimensions[hdr_row].height = 24

    for ri, row in enumerate(rows, start=hdr_row + 1):
        even = (ri - hdr_row) % 2 == 0
        for ci, c in enumerate(cols, start=1):
            fmt = c.get("fmt")
            val = _coerce(row.get(c["key"]), fmt)
            cell = ws.cell(row=ri, column=ci, value=val)
            align = "right" if c.get("align") == "right" else "center" if c.get("align") == "center" else "left"
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.fill = zebra if even else white
            if fmt in _NUMFMT and isinstance(val, (int, float)):
                cell.number_format = _NUMFMT[fmt]
            cell.font = Font(bold=ci == 1, color=_INK, size=10)
        ws.row_dimensions[ri].height = 18

    last_row = hdr_row + len(rows)

    # totals row — live SUM over int/money columns, double-rule on top
    totals_row = None
    summable = [i for i, c in enumerate(cols, start=1) if c.get("fmt") in ("int", "money")]
    if rows and summable:
        totals_row = last_row + 1
        tl = ws.cell(row=totals_row, column=1, value="TOTAL")
        tl.font = Font(bold=True, size=10, color=deep)
        tl.alignment = Alignment(horizontal="left", vertical="center")
        for ci in range(1, ncols + 1):
            cell = ws.cell(row=totals_row, column=ci)
            cell.fill = PatternFill("solid", fgColor=soft)
            cell.border = Border(top=Side(style="double", color=deep), left=thin, right=thin)
            if ci in summable:
                L = get_column_letter(ci)
                cell.value = f"=SUM({L}{hdr_row + 1}:{L}{last_row})"
                cell.number_format = _NUMFMT[cols[ci - 1]["fmt"]]
                cell.font = Font(bold=True, size=10, color=deep)
                cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[totals_row].height = 20

    # borders: full grid over header+data, medium outline around the whole table
    if rows:
        _grid(ws, hdr_row, 1, last_row, ncols, thin)
        _outline(ws, hdr_row, 1, last_row, ncols, med)
        if totals_row:
            _outline(ws, hdr_row, 1, totals_row, ncols, med)
        ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)
        ws.auto_filter.ref = f"A{hdr_row}:{last_col_letter}{last_row}"
    else:
        _grid(ws, hdr_row, 1, hdr_row, ncols, thin)
        _outline(ws, hdr_row, 1, hdr_row, ncols, med)

    # widths
    for i, c in enumerate(cols, start=1):
        L = get_column_letter(i)
        maxlen = max([len(str(c["label"]))] + [len(str(r.get(c["key"]) or "")) for r in rows[:200]] + [8])
        ws.column_dimensions[L].width = min(max(maxlen + 3, 11), 42)

    # traffic-light conditional formatting on rate / gap columns
    for i, c in enumerate(cols, start=1):
        L = get_column_letter(i); rng = f"{L}{hdr_row + 1}:{L}{last_row}"
        if not rows:
            break
        if c.get("fmt") == "pct" and c["key"] in ("completion_rate", "coverage", "pass_rate", "fulfil_rate", "utilization"):
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="num", start_value=0, start_color="FEE2E2",
                mid_type="num", mid_value=60, mid_color="FEF3C7",
                end_type="num", end_value=100, end_color="CCFBF1"))
        elif c["key"] in ("avg_gap", "gap", "overdue", "expired"):
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="greaterThan", formula=["1.5" if c["key"] == "avg_gap" else "0"],
                fill=PatternFill("solid", fgColor="FEE2E2"), font=Font(color="991B1B", bold=True)))

    # native bar chart for the headline metric (guarded — never break export)
    try:
        chart_col = next((i for i, c in enumerate(cols, start=1)
                          if c.get("fmt") == "pct" or c["key"] in ("avg_gap", "avg_rating", "rating", "utilization")), None)
        if chart_col and rows and len(rows) <= 40:
            n = min(len(rows), 20)
            chart = BarChart(); chart.type = "bar"; chart.style = 10
            chart.title = cols[chart_col - 1]["label"]
            chart.height = max(7, n * 0.5); chart.width = 16
            chart.legend = None
            data = Reference(ws, min_col=chart_col, min_row=hdr_row, max_row=hdr_row + n)
            cats = Reference(ws, min_col=1, min_row=hdr_row + 1, max_row=hdr_row + n)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            anchor_col = get_column_letter(len(cols) + 2)
            ws.add_chart(chart, f"{anchor_col}{hdr_row}")
    except Exception:
        pass

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

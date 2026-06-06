"""Excel workbook — PF ECR (Electronic Challan-cum-Return), openpyxl engine.

An ECR-style upload sheet: an EPFO-green title rail, a remittance KPI strip,
a monospace member-row grid (UAN / EPF / EPS / EDLI wage bases + EE/ER split +
NCP days) that mirrors the EPFO Unified Portal layout, a remittance summary
block (EE share / ER EPS / ER EPF / EDLI / admin — totals), traffic-light
heat-map on the employee-PF column, frozen header + autofilter.

Pure openpyxl — chosen for the form-feel, formula-driven summary block and
cell-level traffic-light formatting that suits a statutory upload sheet.
"""
from __future__ import annotations

import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

from ..data import report_meta
from ..common import COMPANY

ACCENT = "15803D"
SOFT = "DCFCE7"
DEEP = "14532D"
INK = "1A1410"
MUTED = "6B5840"
CREAM = "FBF8F0"
MONEY_FMT = '"₹"#,##0'
WHITE = "FFFFFF"


# column spec: (header, row-key, kind)  kind ∈ text|mono|money|days
COLUMNS = [
    ("UAN", "uan", "mono"),
    ("Member ID", "employee_code", "mono"),
    ("Member Name", "employee_name", "text"),
    ("Gross Wages", "gross_wages", "money"),
    ("EPF Wages", "epf_wages", "money"),
    ("EPS Wages", "eps_wages", "money"),
    ("EDLI Wages", "edli_wages", "money"),
    ("EE Share (12%)", "ee_pf", "money"),
    ("ER EPS (8.33%)", "er_eps", "money"),
    ("ER EPF (3.67%)", "er_epf", "money"),
    ("NCP Days", "ncp_days", "days"),
    ("Refund", "refund", "money"),
]
MONEY_KEYS = {"gross_wages", "epf_wages", "eps_wages", "edli_wages",
              "ee_pf", "er_eps", "er_epf", "refund"}


def _thin(color: str) -> Side:
    return Side(style="thin", color=color)


def render(rows: list[dict], summary: dict, meta: dict) -> bytes:
    theme = report_meta("pf-ecr")
    period = meta.get("period", {}) or {}
    label = period.get("label", "")
    fy = period.get("fy", "")
    yr = period.get("year", "")
    mo = period.get("month", 0)
    wage_month = f"{mo:02d}/{yr}" if isinstance(mo, int) and mo else label

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PF ECR"
    ws.sheet_properties.tabColor = ACCENT
    ws.sheet_view.showGridLines = False

    ncol = len(COLUMNS)
    last_letter = get_column_letter(ncol)

    # ── column widths ───────────────────────────────────────────────────────
    widths = [20, 13, 26, 14, 13, 13, 13, 15, 15, 15, 10, 11]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    thin_a = _thin(ACCENT + "55") if False else Side(style="thin", color="BBD8C2")
    box = Border(left=thin_a, right=thin_a, top=thin_a, bottom=thin_a)

    r = 1
    # ── accent title rail (height ~4) ────────────────────────────────────────
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor=ACCENT)
    ws.row_dimensions[r].height = 5
    r += 1

    # ── masthead title ───────────────────────────────────────────────────────
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    c = ws.cell(r, 1, f"  {COMPANY['name']}  ·  EMPLOYEES' PROVIDENT FUND ORGANISATION")
    c.font = Font(name="Calibri", size=8, bold=True, color=DEEP)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 16
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    c = ws.cell(r, 1, "  Electronic Challan-cum-Return (ECR)")
    c.font = Font(name="Calibri", size=18, bold=True, color=INK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 30
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    c = ws.cell(r, 1, "  EPFO upload format — UAN · wages · EE/ER split · NCP days")
    c.font = Font(name="Calibri", size=10, italic=True, color=MUTED)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = Border(bottom=Side(style="thin", color="CBD5E1"))
    ws.row_dimensions[r].height = 18
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    c = ws.cell(r, 1, f"  Wage month  {wage_month}  ·  {label}      ·      FY {fy}"
                      f"      ·      Establishment  {COMPANY['legal']}  ·  TAN {COMPANY['tan']}")
    c.font = Font(name="Calibri", size=9, bold=True, color=INK)
    c.fill = PatternFill("solid", fgColor="F1F5E9")
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = Border(top=Side(style="thin", color=ACCENT),
                      bottom=Side(style="medium", color=DEEP))
    ws.row_dimensions[r].height = 20
    r += 2

    # ── KPI tiles (merged, colored thick top border) ─────────────────────────
    members = summary.get("rows", len(rows))
    employees = summary.get("employees", members)
    pf_total = summary.get("pf", 0.0)
    gross_total = summary.get("gross", 0.0)
    ee_total = sum(float(x.get("ee_pf") or 0) for x in rows)
    er_total = pf_total - ee_total if pf_total else sum(
        float(x.get("er_eps") or 0) + float(x.get("er_epf") or 0) for x in rows)

    kpis = [
        ("MEMBERS", members, ACCENT, None),
        ("UAN RECORDS", employees, "475569", None),
        ("EE SHARE", ee_total, ACCENT, MONEY_FMT),
        ("ER SHARE", er_total, "B45309", MONEY_FMT),
        ("TOTAL EPF PAYABLE", pf_total, DEEP, MONEY_FMT),
    ]
    nk = len(kpis)
    span = ncol // nk
    kpi_label_row = r
    kpi_value_row = r + 1
    ws.row_dimensions[kpi_label_row].height = 16
    ws.row_dimensions[kpi_value_row].height = 26
    col = 1
    for i, (lbl, val, rail, fmt) in enumerate(kpis):
        extra = 1 if i < (ncol - span * nk) else 0
        c1 = col
        c2 = min(col + span - 1 + extra, ncol)
        ws.merge_cells(start_row=kpi_label_row, start_column=c1, end_row=kpi_label_row, end_column=c2)
        lc = ws.cell(kpi_label_row, c1, lbl)
        lc.font = Font(name="Calibri", size=7.5, bold=True, color="475569")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.fill = PatternFill("solid", fgColor=WHITE)
        lc.border = Border(top=Side(style="thick", color=rail),
                           left=Side(style="thin", color="CBD5E1"),
                           right=Side(style="thin", color="CBD5E1"))
        ws.merge_cells(start_row=kpi_value_row, start_column=c1, end_row=kpi_value_row, end_column=c2)
        vc = ws.cell(kpi_value_row, c1, val)
        vc.font = Font(name="Calibri", size=16, bold=True, color=INK)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.fill = PatternFill("solid", fgColor=WHITE)
        vc.border = Border(left=Side(style="thin", color="CBD5E1"),
                           right=Side(style="thin", color="CBD5E1"),
                           bottom=Side(style="thin", color="CBD5E1"))
        if fmt:
            vc.number_format = fmt
        col = c2 + 1
    r = kpi_value_row + 2

    # ── themed header row ────────────────────────────────────────────────────
    header_row = r
    head_fill = PatternFill("solid", fgColor=ACCENT)
    head_font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    head_border = Border(top=Side(style="medium", color=DEEP),
                         bottom=Side(style="medium", color=DEEP),
                         left=Side(style="thin", color=DEEP),
                         right=Side(style="thin", color=DEEP))
    for i, (hdr, _key, kind) in enumerate(COLUMNS, start=1):
        c = ws.cell(header_row, i, hdr)
        c.fill = head_fill
        c.font = head_font
        c.border = head_border
        align = "left" if kind in ("text", "mono") else "right"
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    ws.row_dimensions[header_row].height = 26
    ws.freeze_panes = ws.cell(header_row + 1, 1)

    # ── body rows (zebra, monospace refs) ────────────────────────────────────
    first_data = header_row + 1
    for ri, row in enumerate(rows):
        rr = first_data + ri
        zebra = ri % 2 == 1
        bg = CREAM if zebra else WHITE
        for i, (_hdr, key, kind) in enumerate(COLUMNS, start=1):
            v = row.get(key)
            c = ws.cell(rr, i)
            c.fill = PatternFill("solid", fgColor=bg)
            c.border = box
            if kind == "money":
                c.value = float(v or 0)
                c.number_format = MONEY_FMT
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.font = Font(name="Calibri", size=10, color=INK)
            elif kind == "days":
                c.value = float(v or 0)
                c.number_format = "0.0"
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.font = Font(name="Calibri", size=10, color=INK)
            elif kind == "mono":
                c.value = v if v not in (None, "") else "—"
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.font = Font(name="Consolas", size=9.5, color=INK)
            else:
                c.value = v if v not in (None, "") else "—"
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.font = Font(name="Calibri", size=10, color=INK)
    last_data = first_data + len(rows) - 1 if rows else header_row

    # ── remittance TOTAL row (formula-summed money columns) ──────────────────
    total_row = last_data + 1 if rows else header_row + 1
    tot_fill = PatternFill("solid", fgColor=DEEP)
    tot_font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    tot_border = Border(top=Side(style="medium", color=DEEP),
                        bottom=Side(style="medium", color=DEEP),
                        left=Side(style="thin", color=DEEP),
                        right=Side(style="thin", color=DEEP))
    for i, (_hdr, key, kind) in enumerate(COLUMNS, start=1):
        c = ws.cell(total_row, i)
        c.fill = tot_fill
        c.border = tot_border
        if i == 1:
            c.value = "TOTAL"
            c.font = tot_font
            c.alignment = Alignment(horizontal="left", vertical="center")
        elif i == 3:
            c.value = f"{len(rows)} members"
            c.font = tot_font
            c.alignment = Alignment(horizontal="left", vertical="center")
        elif key in MONEY_KEYS and rows:
            cl = get_column_letter(i)
            c.value = f"=SUM({cl}{first_data}:{cl}{last_data})"
            c.font = tot_font
            c.number_format = MONEY_FMT
            c.alignment = Alignment(horizontal="right", vertical="center")
        elif kind == "days" and rows:
            cl = get_column_letter(i)
            c.value = f"=SUM({cl}{first_data}:{cl}{last_data})"
            c.font = tot_font
            c.number_format = "0.0"
            c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[total_row].height = 22

    # ── remittance summary block (challan breakup) ───────────────────────────
    sb = total_row + 2
    ws.merge_cells(start_row=sb, start_column=1, end_row=sb, end_column=4)
    c = ws.cell(sb, 1, "  REMITTANCE SUMMARY — CHALLAN BREAK-UP")
    c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=ACCENT)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[sb].height = 20

    # break-up uses formulas referencing the total row where possible
    def _col(key):
        for i, (_h, k, _kind) in enumerate(COLUMNS, start=1):
            if k == key:
                return get_column_letter(i)
        return "A"

    if rows:
        ee_ref = f"{_col('ee_pf')}{total_row}"
        eps_ref = f"{_col('er_eps')}{total_row}"
        epf_ref = f"{_col('er_epf')}{total_row}"
        edli_base = f"{_col('edli_wages')}{total_row}"
        breakup = [
            ("A/C 01 — EPF (EE 12% + ER 3.67%)", f"={ee_ref}+{epf_ref}"),
            ("A/C 10 — EPS (Employer 8.33%)", f"={eps_ref}"),
            ("A/C 21 — EDLI (0.50% of EDLI wages)", f"=ROUND({edli_base}*0.005,0)"),
            ("A/C 02 — EPF Admin (0.50%)", f"=ROUND({_col('epf_wages')}{total_row}*0.005,0)"),
            ("A/C 22 — EDLI Admin", 0),
        ]
    else:
        breakup = [
            ("A/C 01 — EPF (EE 12% + ER 3.67%)", 0),
            ("A/C 10 — EPS (Employer 8.33%)", 0),
            ("A/C 21 — EDLI (0.50% of EDLI wages)", 0),
            ("A/C 02 — EPF Admin (0.50%)", 0),
            ("A/C 22 — EDLI Admin", 0),
        ]

    br = sb + 1
    sum_box = Border(left=_thin("BBD8C2"), right=_thin("BBD8C2"),
                     top=_thin("BBD8C2"), bottom=_thin("BBD8C2"))
    for idx, (lbl, val) in enumerate(breakup):
        bg = CREAM if idx % 2 else WHITE
        ws.merge_cells(start_row=br, start_column=1, end_row=br, end_column=3)
        lc = ws.cell(br, 1, lbl)
        lc.font = Font(name="Calibri", size=9.5, color=INK)
        lc.fill = PatternFill("solid", fgColor=bg)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lc.border = sum_box
        ws.cell(br, 2).border = sum_box
        ws.cell(br, 3).border = sum_box
        vc = ws.cell(br, 4, val)
        vc.number_format = MONEY_FMT
        vc.font = Font(name="Calibri", size=10, bold=True, color=DEEP)
        vc.fill = PatternFill("solid", fgColor=bg)
        vc.alignment = Alignment(horizontal="right", vertical="center")
        vc.border = sum_box
        br += 1

    # grand total of challan
    ws.merge_cells(start_row=br, start_column=1, end_row=br, end_column=3)
    gc = ws.cell(br, 1, "  TOTAL CHALLAN AMOUNT (Rounded)")
    gc.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    gc.fill = PatternFill("solid", fgColor=DEEP)
    gc.alignment = Alignment(horizontal="left", vertical="center")
    gv_col = get_column_letter(4)
    gv = ws.cell(br, 4, f"=SUM({gv_col}{sb + 1}:{gv_col}{br - 1})" if True else 0)
    gv.number_format = MONEY_FMT
    gv.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    gv.fill = PatternFill("solid", fgColor=DEEP)
    gv.alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(br, 2).fill = PatternFill("solid", fgColor=DEEP)
    ws.cell(br, 3).fill = PatternFill("solid", fgColor=DEEP)
    ws.row_dimensions[br].height = 22

    # ── autofilter + traffic-light heat-map on EE Share column ───────────────
    if rows:
        ws.auto_filter.ref = f"A{header_row}:{last_letter}{last_data}"
        ee_letter = _col("ee_pf")
        rng = f"{ee_letter}{first_data}:{ee_letter}{last_data}"
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="min", start_color=SOFT,
                mid_type="percentile", mid_value=50, mid_color="86EFAC",
                end_type="max", end_color=ACCENT,
            ),
        )

    # print niceties
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

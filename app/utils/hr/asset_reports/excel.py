"""HR Asset Reports — Excel exporter (openpyxl, dashboard-grade workbook).

Sheet 1 "Dashboard": an accent title band, a row of bordered KPI cards (accent
rail + label + big value) and a native bar chart of the report's headline
distribution. Sheet 2 "Data": a fully-bordered, banded table with per-column
number formats, a SUM totals row, frozen header, autofilter and conditional
formatting (colour-scale on %, red flags on risk columns). Generic over the
report shape via the column ``fmt``/``align`` descriptors.
"""
from __future__ import annotations

import io
from datetime import date, datetime

from .data import report_meta

_NUMFMT = {"money": '₹#,##0', "int": '#,##0', "pct": '0.0"%"'}
_RISK = {  # column key -> (operator, threshold) for a red CellIs flag
    "overdue": ("greaterThan", 0), "missing": ("greaterThan", 0), "mismatched": ("greaterThan", 0),
    "over_7d": ("greaterThan", 0), "days_pending": ("greaterThan", 7), "age_months": ("greaterThanOrEqual", 36),
    "days_left": ("lessThan", 0),
}
_SUMMABLE = {"money", "int"}


def _hex(c):
    return (c or "#fbbf24").lstrip("#").upper()


def _coerce(value, fmt):
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (datetime, date)):
        return value
    if fmt in _NUMFMT and isinstance(value, (int, float)):
        return value
    return value


def _chart_series(report):
    """Pick a headline distribution to chart: prefer a nested-dict summary
    (by_status / by_method …), else the top rows by the first numeric column."""
    summary = report.get("summary", {})
    for key, val in summary.items():
        if isinstance(val, dict) and val:
            title = key.replace("by_", "").replace("_", " ").title()
            pairs = [(str(k).replace("_", " ").title(), float(v or 0)) for k, v in val.items()]
            return title, pairs[:10]
    cols = report["columns"]
    rows = report["rows"]
    num_col = next((c for c in cols[1:] if c.get("fmt") in ("int", "money")), None)
    if not num_col or not rows:
        return None, []
    label_key = cols[0]["key"]
    pairs = [(str(r.get(label_key, "")), float(r.get(num_col["key"]) or 0)) for r in rows[:10]]
    return num_col["label"], pairs


def render_excel(report: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.shapes import GraphicalProperties

    meta = report_meta(report["key"])
    accent = _hex(meta.get("accent", "#fbbf24"))
    deep = _hex(meta.get("accent_deep", "#b45309"))
    soft = _hex(meta.get("accent_soft", "#fff7e6"))

    INK, MUT = "1A1410", "6B5840"
    thin = Side(style="thin", color="E5DCCB")
    med = Side(style="medium", color=deep)
    deep_side = Side(style="thin", color=deep)
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    # ════════════ Sheet 1 — Dashboard ════════════
    ds = wb.active
    ds.title = "Dashboard"
    ds.sheet_view.showGridLines = False
    ds.sheet_properties.tabColor = accent
    ds.column_dimensions["A"].width = 2
    for col in "BCEFHIKL":
        ds.column_dimensions[col].width = 11.5
    for col in "DGJ":
        ds.column_dimensions[col].width = 2.4

    ds.merge_cells("B2:L2")
    t = ds["B2"]
    t.value = report.get("title", "Report")
    t.font = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for cl in "BCDEFGHIJKL":
        ds[f"{cl}2"].fill = PatternFill("solid", fgColor=deep)
    ds.row_dimensions[2].height = 34

    ds.merge_cells("B3:L3")
    ds["B3"].value = report.get("subtitle", "")
    ds["B3"].font = Font(size=11, color=MUT)
    ds["B3"].alignment = Alignment(indent=1)
    ds.merge_cells("B4:L4")
    ds["B4"].value = (f'{meta.get("eyebrow", "ASSET HANGAR")}   ·   '
                      f'Period: {report.get("period", {}).get("label", "All time")}   ·   '
                      f'Generated {date.today().isoformat()}')
    ds["B4"].font = Font(size=8.5, bold=True, color=accent)
    ds["B4"].alignment = Alignment(indent=1)

    # KPI cards — 4 per band
    summary = report.get("summary", {})
    tiles = []
    for k, v in summary.items():
        if isinstance(v, (dict, bool)) or v is None:
            continue
        label = k.replace("_", " ").title()
        if any(tok in k for tok in ("value", "cost", "book", "sale", "spend", "recovery")):
            val = f"₹{float(v):,.0f}" if isinstance(v, (int, float)) else str(v)
        elif "pct" in k or "accuracy" in k or "rate" in k:
            val = f"{float(v):g}%" if isinstance(v, (int, float)) else str(v)
        else:
            val = f"{int(v):,}" if isinstance(v, (int, float)) and float(v).is_integer() else str(v)
        tiles.append((label, val))

    cols4 = [(2, 3), (5, 6), (8, 9), (11, 12)]
    r0 = 6
    shown = tiles[:8]
    for i, (label, val) in enumerate(shown):
        band = i // 4
        lc, rc = cols4[i % 4]
        top = r0 + band * 5
        ds.merge_cells(start_row=top, start_column=lc, end_row=top, end_column=rc)
        for c in (lc, rc):
            ds.cell(top, c).fill = PatternFill("solid", fgColor=accent)
        ds.row_dimensions[top].height = 5
        ds.merge_cells(start_row=top + 1, start_column=lc, end_row=top + 1, end_column=rc)
        lab = ds.cell(top + 1, lc, label)
        lab.font = Font(size=8, bold=True, color=MUT)
        lab.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ds.merge_cells(start_row=top + 2, start_column=lc, end_row=top + 2, end_column=rc)
        vc = ds.cell(top + 2, lc, val)
        vc.font = Font(size=15, bold=True, color=deep)
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ds.row_dimensions[top + 2].height = 26
        for rr in (top + 1, top + 2):
            for c in (lc, rc):
                ds.cell(rr, c).fill = PatternFill("solid", fgColor=soft)
        for c in (lc, rc):
            for rr in (top, top + 1, top + 2):
                ds.cell(rr, c).border = Border(
                    left=thin if c == lc else None, right=thin if c == rc else None,
                    top=thin if rr == top else None, bottom=thin if rr == top + 2 else None)

    # chart
    ctitle, series = _chart_series(report)
    bands = (len(shown) + 3) // 4 if shown else 1
    chart_row = r0 + bands * 5 + 1
    if series:
        hcol = 15  # column O (hidden helper block)
        ds.cell(1, hcol, "label")
        ds.cell(1, hcol + 1, "value")
        for j, (lab, val) in enumerate(series, start=2):
            ds.cell(j, hcol, lab)
            ds.cell(j, hcol + 1, val)
        ds.column_dimensions[get_column_letter(hcol)].hidden = True
        ds.column_dimensions[get_column_letter(hcol + 1)].hidden = True

        chart = BarChart()
        chart.type = "col"
        chart.title = ctitle or "Distribution"
        chart.legend = None
        chart.height = 7.4
        chart.width = 20
        data = Reference(ds, min_col=hcol + 1, min_row=1, max_row=len(series) + 1)
        cats = Reference(ds, min_col=hcol, min_row=2, max_row=len(series) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties = GraphicalProperties(solidFill=accent)
        ds.add_chart(chart, f"B{chart_row}")

    # ════════════ Sheet 2 — Data ════════════
    dt = wb.create_sheet("Data")
    dt.sheet_view.showGridLines = False
    dt.sheet_properties.tabColor = deep
    cols = report["columns"]
    head_fill = PatternFill("solid", fgColor=deep)
    head_font = Font(bold=True, color="FFFFFF", size=9.5)
    band_fill = PatternFill("solid", fgColor=soft)
    head_border = Border(left=deep_side, right=deep_side, top=deep_side, bottom=deep_side)

    for ci, c in enumerate(cols, start=1):
        cell = dt.cell(1, ci, c["label"])
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal=("right" if c.get("align") == "right" else "left"), vertical="center")
        cell.border = head_border
        dt.column_dimensions[get_column_letter(ci)].width = max(11, min(34, len(c["label"]) + 6))
    dt.row_dimensions[1].height = 22

    n = len(report["rows"])
    for ri, row in enumerate(report["rows"], start=2):
        for ci, c in enumerate(cols, start=1):
            cell = dt.cell(ri, ci, _coerce(row.get(c["key"]), c.get("fmt")))
            cell.font = Font(size=10, color=INK)
            cell.border = box
            if c.get("fmt") in _NUMFMT:
                cell.number_format = _NUMFMT[c["fmt"]]
            if c.get("align") == "right":
                cell.alignment = Alignment(horizontal="right")
            if ri % 2 == 0:
                cell.fill = band_fill

    if n:
        tr = n + 2
        for ci, c in enumerate(cols, start=1):
            cell = dt.cell(tr, ci)
            cell.fill = PatternFill("solid", fgColor=soft)
            cell.border = Border(top=med, bottom=med)
            cell.font = Font(bold=True, color=deep, size=10)
            if ci == 1:
                cell.value = "TOTAL"
            elif c.get("fmt") in _SUMMABLE:
                col = get_column_letter(ci)
                cell.value = f"=SUM({col}2:{col}{n + 1})"
                cell.number_format = _NUMFMT[c["fmt"]]
                cell.alignment = Alignment(horizontal="right")

        last = get_column_letter(len(cols))
        dt.auto_filter.ref = f"A1:{last}{n + 1}"
        dt.freeze_panes = "A2"

        for ci, c in enumerate(cols, start=1):
            col = get_column_letter(ci)
            rng = f"{col}2:{col}{n + 1}"
            key = c["key"]
            if c.get("fmt") == "pct":
                dt.conditional_formatting.add(rng, ColorScaleRule(
                    start_type="num", start_value=0, start_color="FEE2E2",
                    mid_type="num", mid_value=50, mid_color="FEF3C7",
                    end_type="num", end_value=100, end_color="DCFCE7"))
            elif key in _RISK:
                op, thr = _RISK[key]
                dt.conditional_formatting.add(rng, CellIsRule(
                    operator=op, formula=[str(thr)],
                    fill=PatternFill("solid", fgColor="FEE2E2"), font=Font(color="991B1B", bold=True)))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

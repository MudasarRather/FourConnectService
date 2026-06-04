"""Excel exports for HR Attendance Reports.

Seven reports, seven layouts. We use both libraries:

    * ``xlsxwriter`` — for the four reports that need native charts or
      sparklines (overtime gauges, daily roster pivot, late heat-map,
      monthly summary bars). xlsxwriter writes a fresh workbook with rich
      conditional formats.
    * ``openpyxl`` — for the three reports that benefit from formula
      power and a more spreadsheet-native feel (compliance audit with
      traffic-light formatting, anomalies dossier with row tints, WFH
      remote-work journal with merged headers).

Each renderer takes the shaped row list and the summary dict and returns
``bytes`` (the .xlsx blob) so the router can wrap it in a Response.

Design philosophy:
    * Frozen header row on every sheet.
    * Autofilter on the main table.
    * One title band per report, themed to match the PDF cover accent.
    * Conditional formats highlight outliers without burying them.
    * Corporate typography — Calibri / Aptos at restrained sizes (18-22pt
      titles, 9-10pt body). No editorial-style display fonts that look
      out of place in a spreadsheet.
"""
from __future__ import annotations

import io
from datetime import datetime, date as date_cls

# Lazy-imported in render() so missing deps fail loudly only at export time.

from .data import report_meta, STATUS_COLORS


COMPANY = {
    "name": "Fourreck",
    "legal": "Fourreck Technologies Pvt. Ltd.",
    "web": "fourreck.com",
}


# ════════════════════════════════════════════════════════════════════════════
# Shared corporate palette — unified neutrals across every report so titles,
# subtitles, period bands and zebra rows feel like one product line. Each
# report still gets its own accent color (drives the KPI top-rail, tab tint,
# header row + chart fills) so visual identity per report is preserved.
# ════════════════════════════════════════════════════════════════════════════
BRAND = {
    "ink":          "#111418",   # primary text
    "ink_muted":    "#475569",   # secondary text
    "ink_dim":      "#94a3b8",   # tertiary
    "rule":         "#94a3b8",   # body cell borders — strong enough to read on white
    "rule_soft":    "#cbd5e1",   # zebra row dividers / column rules
    "rule_strong":  "#475569",   # header rules + period band underline
    "cream":        "#fbf8f0",   # zebra fill (warm cream)
    "panel":        "#ffffff",   # KPI tile background
    "panel_soft":   "#f1f5f9",   # period band background
    "header_ink":   "#ffffff",   # header text on accent
    "danger_bg":    "#fee2e2", "danger_fg": "#7f1d1d",
    "warn_bg":      "#fef3c7", "warn_fg":   "#854d0e",
    "good_bg":      "#ccfbf1", "good_fg":   "#115e59",
    "title_pt":     18,          # title font size
    "subtitle_pt":  10,          # subtitle / tagline
    "meta_pt":      9,           # period · generated line
    "kpi_label_pt": 8,
    "kpi_value_pt": 18,
    "header_pt":    10,
    "body_pt":      10,
}


def _corporate_title_block(wb, ws, theme: dict, period: dict, summary: dict, *, last_col: int) -> int:
    """Write a uniform 4-row corporate header: top accent rail · title · subtitle · period.

    Returns the row index immediately after the block (caller should start KPI
    or data table from there).

        last_col: zero-based last column index the block should span across.
    """
    accent = theme["accent"]
    deep = theme["accent_deep"]
    name = theme["name"]
    subtitle = theme.get("subtitle", "")

    f_rail = wb.add_format({"bg_color": accent})
    f_title = wb.add_format({
        "bold": True, "font_size": BRAND["title_pt"], "font_name": "Calibri",
        "font_color": BRAND["ink"], "bg_color": BRAND["panel"],
        "align": "left", "valign": "vcenter", "indent": 1,
    })
    f_sub = wb.add_format({
        "font_size": BRAND["subtitle_pt"], "italic": True, "font_name": "Calibri",
        "font_color": BRAND["ink_muted"], "bg_color": BRAND["panel"],
        "align": "left", "valign": "vcenter", "indent": 1,
        "bottom": 1, "bottom_color": BRAND["rule_soft"],
    })
    f_period = wb.add_format({
        "bold": True, "font_size": BRAND["meta_pt"], "font_name": "Calibri",
        "font_color": BRAND["ink"], "bg_color": BRAND["panel_soft"],
        "align": "left", "valign": "vcenter", "indent": 1,
        "top": 1, "top_color": BRAND["rule"],
        "bottom": 2, "bottom_color": deep,
    })

    ws.set_row(0, 4)                           # accent rail
    ws.merge_range(0, 0, 0, last_col, "", f_rail)

    ws.set_row(1, 32)                          # title
    ws.merge_range(
        1, 0, 1, last_col,
        f"  {COMPANY['name']}  ·  {name}",
        f_title,
    )

    ws.set_row(2, 20)                          # subtitle
    ws.merge_range(2, 0, 2, last_col, f"  {subtitle}", f_sub)

    ws.set_row(3, 22)                          # period · generated
    ws.merge_range(
        3, 0, 3, last_col,
        f"  Period   {period['from'].strftime('%d %b %Y')}    →    {period['to'].strftime('%d %b %Y')}"
        f"        ·        Generated   {datetime.now().strftime('%d %b %Y · %I:%M %p').lstrip('0')}",
        f_period,
    )

    return 4


def _corporate_kpi_strip(wb, ws, theme: dict, kpis: list[tuple[str, object, str]], *,
                          start_row: int, last_col: int) -> int:
    """Write a clean KPI row: label row (8pt) + value row (16pt) with a
    1-cell-wide colored top accent on each tile.

        kpis: list of (LABEL, value, hex_color_for_top_rail).
            value can be int / float / str — str values render as-is.
        Returns the row index immediately after the strip.
    """
    if not kpis or last_col < 0:
        return start_row

    n = len(kpis)
    # We span the full available columns evenly across n tiles.
    cols_per = max(1, (last_col + 1) // n)
    leftover = (last_col + 1) - cols_per * n

    # Spacer row above
    ws.set_row(start_row, 8)

    label_row = start_row + 1
    value_row = start_row + 2
    ws.set_row(label_row, 20)
    ws.set_row(value_row, 30)

    c0 = 0
    for i, (label, value, rail_color) in enumerate(kpis):
        extra = 1 if i < leftover else 0
        c1 = min(c0 + cols_per - 1 + extra, last_col)

        f_label = wb.add_format({
            "bold": True, "font_size": BRAND["kpi_label_pt"], "font_name": "Calibri",
            "font_color": BRAND["ink_muted"], "bg_color": BRAND["panel"],
            "align": "center", "valign": "vcenter",
            "left": 2, "left_color": BRAND["rule"],
            "right": 2, "right_color": BRAND["rule"],
            "top": 5, "top_color": rail_color,            # 5 = "thick" → visible top rail
        })
        f_value = wb.add_format({
            "bold": True, "font_size": BRAND["kpi_value_pt"], "font_name": "Calibri",
            "font_color": BRAND["ink"], "bg_color": BRAND["panel"],
            "align": "center", "valign": "vcenter",
            "left": 2, "left_color": BRAND["rule"],
            "right": 2, "right_color": BRAND["rule"],
            "bottom": 2, "bottom_color": BRAND["rule"],
        })

        if c0 == c1:
            ws.write(label_row, c0, label, f_label)
            if isinstance(value, (int, float)):
                ws.write_number(value_row, c0, value, f_value)
            else:
                ws.write(value_row, c0, str(value), f_value)
        else:
            ws.merge_range(label_row, c0, label_row, c1, label, f_label)
            if isinstance(value, (int, float)):
                # merge first, then overwrite center cell with the number
                ws.merge_range(value_row, c0, value_row, c1, "", f_value)
                ws.write_number(value_row, c0, value, f_value)
            else:
                ws.merge_range(value_row, c0, value_row, c1, str(value), f_value)
        c0 = c1 + 1

    # Trailing spacer
    spacer = value_row + 1
    ws.set_row(spacer, 10)
    return spacer + 1


def _corporate_header_format(wb, theme: dict, *, align: str = "left") -> object:
    """Reusable table-header format. Accent fill, white text, dark bottom rule
    + thin side rules so the header reads as a discrete band over the body grid.
    """
    return wb.add_format({
        "bold": True, "font_color": BRAND["header_ink"],
        "bg_color": theme["accent"], "font_name": "Calibri",
        "align": align, "valign": "vcenter", "indent": 1,
        "font_size": BRAND["header_pt"],
        "top": 2, "top_color": theme["accent_deep"],
        "bottom": 2, "bottom_color": theme["accent_deep"],
        "left": 1, "left_color": theme["accent_deep"],
        "right": 1, "right_color": theme["accent_deep"],
        "text_wrap": False,
    })


# ════════════════════════════════════════════════════════════════════════════
# Dispatch
# ════════════════════════════════════════════════════════════════════════════


def render_excel(report_key: str, shaped_rows: list[dict], summary: dict, meta: dict) -> bytes:
    fn = RENDERERS.get(report_key, _excel_daily)
    return fn(shaped_rows, summary, meta)


# ════════════════════════════════════════════════════════════════════════════
# xlsxwriter helpers
# ════════════════════════════════════════════════════════════════════════════


def _xw_workbook():
    """Construct an xlsxwriter workbook backed by an in-memory buffer.

    ``remove_timezone=True`` is essential — our check_in_time / check_out_time
    columns come back from PostgreSQL as IST-aware datetimes, but Excel has no
    timezone concept so xlsxwriter raises TypeError on tz-aware input. The
    flag strips tzinfo at write time (preserving the wall-clock value, which
    is what HR cares about anyway).
    """
    import xlsxwriter
    buf = io.BytesIO()
    wb = xlsxwriter.Workbook(buf, {
        "in_memory": True,
        "default_date_format": "yyyy-mm-dd",
        "remove_timezone": True,
    })
    return wb, buf


def _xw_finalize(wb, buf) -> bytes:
    wb.close()
    buf.seek(0)
    return buf.read()


def _fmt_time(v):
    if isinstance(v, datetime):
        return v.strftime("%I:%M %p").lstrip("0")
    if v is None:
        return ""
    return str(v)


def _coverage_pct(v):
    """Force coverage_pct values to be plain numbers for conditional formatting."""
    try:
        return float(v) / 100.0
    except (TypeError, ValueError):
        return 0


# ════════════════════════════════════════════════════════════════════════════
# 1. Monthly Summary — xlsxwriter with embedded BAR CHART
# ════════════════════════════════════════════════════════════════════════════


def _excel_monthly(rows: list[dict], summary: dict, meta: dict) -> bytes:
    theme = report_meta("monthly")
    accent = theme["accent"]
    accent_soft = theme["accent_soft"]
    deep = theme["accent_deep"]
    period = meta["period"]

    wb, buf = _xw_workbook()
    ws = wb.add_worksheet("Monthly Summary")
    ws.set_tab_color(accent)
    ws.hide_gridlines(2)

    # ── Layout ─────────────────────────────────────────────────────────
    ws.set_column("A:A", 12)   # code
    ws.set_column("B:B", 24)   # name
    ws.set_column("C:C", 18)   # department
    ws.set_column("D:D", 16)   # designation
    ws.set_column("E:E", 16)   # shift
    ws.set_column("F:Y", 11)   # numeric payroll columns

    # ── Corporate title block (rows 0..3) ──────────────────────────────
    next_row = _corporate_title_block(wb, ws, theme, period, summary, last_col=11)

    # ── KPI strip — payroll-focused tiles ──────────────────────────────
    present_total = sum(r["present_days"] for r in rows)
    absent_total = sum(r["absent_days"] for r in rows)
    leave_total = round(sum(r["leave_days"] for r in rows), 1)
    lop_total = round(sum(r["lop_days"] for r in rows), 1)
    ot_total = round(sum(r["total_overtime_hours"] for r in rows), 1)
    kpis = [
        ("EMPLOYEES",    summary["employees"], "#475569"),
        ("PRESENT DAYS", present_total,        "#0d9488"),
        ("ABSENT DAYS",  absent_total,         "#b91c1c"),
        ("LEAVE DAYS",   leave_total,          "#7c3aed"),
        ("LOP DAYS",     lop_total,            "#dc2626"),
        ("OVERTIME HRS", ot_total,             "#ea580c"),
    ]
    next_row = _corporate_kpi_strip(wb, ws, theme, kpis, start_row=next_row, last_col=11)

    # ── Body formats ───────────────────────────────────────────────────
    f_header = _corporate_header_format(wb, theme, align="left")
    f_header_r = _corporate_header_format(wb, theme, align="right")
    f_cell = wb.add_format({"font_size": BRAND["body_pt"], "align": "left", "indent": 1, "border": 1, "border_color": BRAND["rule_soft"]})
    f_cell_zebra = wb.add_format({"font_size": BRAND["body_pt"], "align": "left", "indent": 1, "bg_color": BRAND["cream"], "border": 1, "border_color": BRAND["rule_soft"]})
    f_num = wb.add_format({"font_size": BRAND["body_pt"], "align": "right", "indent": 1, "num_format": "0", "border": 1, "border_color": BRAND["rule_soft"]})
    f_num_z = wb.add_format({"font_size": BRAND["body_pt"], "align": "right", "indent": 1, "num_format": "0", "bg_color": BRAND["cream"], "border": 1, "border_color": BRAND["rule_soft"]})
    f_hrs = wb.add_format({"font_size": BRAND["body_pt"], "align": "right", "indent": 1, "num_format": "0.00\" h\"", "border": 1, "border_color": BRAND["rule_soft"]})
    f_hrs_z = wb.add_format({"font_size": BRAND["body_pt"], "align": "right", "indent": 1, "num_format": "0.00\" h\"", "bg_color": BRAND["cream"], "border": 1, "border_color": BRAND["rule_soft"]})

    from xlsxwriter.utility import xl_col_to_name as _xlc

    # ── Full payroll column spec: (label, key, kind) ───────────────────
    COLS = [
        ("Code", "employee_code", "text"),
        ("Employee", "employee_name", "text"),
        ("Department", "department", "text"),
        ("Designation", "designation", "text"),
        ("Shift", "shift_name", "text"),
        ("Present", "present_days", "int"),
        ("Late", "late_days", "int"),
        ("Half", "half_days", "int"),
        ("Absent", "absent_days", "int"),
        ("LWP", "lwp_days", "int"),
        ("Leave", "leave_days", "days"),
        ("WFH", "wfh_days", "int"),
        ("Holiday", "holidays", "int"),
        ("Week-off", "week_offs", "int"),
        ("LOP Days", "lop_days", "days"),
        ("Payable Days", "payable_days", "days"),
        ("Attendance %", "attendance_pct", "pct"),
        ("Working hrs", "total_working_hours", "hours"),
        ("Avg hrs/day", "avg_working_hours", "hours"),
        ("Break hrs", "total_break_hours", "hours"),
        ("Excess Break (min)", "excess_break_minutes", "int"),
        ("OT hrs", "total_overtime_hours", "hours"),
        ("Late (min)", "total_late_minutes", "int"),
        ("Early-exit (min)", "total_early_exit_minutes", "int"),
        ("Flagged", "flagged_days", "int"),
        ("Missing Punch", "missing_punch_days", "int"),
    ]
    ncol = len(COLS)
    col_idx = {k: i for i, (_l, k, _kind) in enumerate(COLS)}

    def _mk(numfmt, zebra=False):
        f = {"font_size": BRAND["body_pt"], "align": "right", "indent": 1,
             "num_format": numfmt, "border": 1, "border_color": BRAND["rule_soft"]}
        if zebra:
            f["bg_color"] = BRAND["cream"]
        return wb.add_format(f)
    f_days, f_days_z = _mk("0.0"), _mk("0.0", True)
    f_pct, f_pct_z = _mk('0.0"%"'), _mk('0.0"%"', True)
    fmt_for = {
        "text":  (f_cell, f_cell_zebra),
        "int":   (f_num, f_num_z),
        "days":  (f_days, f_days_z),
        "pct":   (f_pct, f_pct_z),
        "hours": (f_hrs, f_hrs_z),
    }

    # Header
    start_row = next_row
    for i, (label, _k, kind) in enumerate(COLS):
        ws.write(start_row, i, label, f_header if kind == "text" else f_header_r)
    ws.set_row(start_row, 30)
    ws.freeze_panes(start_row + 1, 3)   # keep header + Code/Employee/Dept on screen

    # Body
    for ri, r in enumerate(rows):
        row_idx = start_row + 1 + ri
        zebra = ri % 2 == 1
        ws.set_row(row_idx, 21)
        for i, (_label, key, kind) in enumerate(COLS):
            base, zeb = fmt_for[kind]
            f = zeb if zebra else base
            v = r.get(key)
            if kind == "text":
                ws.write(row_idx, i, v if v not in (None, "") else "—", f)
            else:
                ws.write_number(row_idx, i, float(v or 0), f)

    last_row = start_row + len(rows)

    # ── TOTAL row (sum count/hours columns; ratio columns left blank) ──
    if rows:
        tr = last_row + 1
        ws.set_row(tr, 24)
        f_tot_l = wb.add_format({"font_size": BRAND["body_pt"], "bold": True, "font_color": "#ffffff",
                                 "bg_color": deep, "align": "left", "indent": 1, "border": 1, "border_color": deep})
        def _tot(numfmt):
            return wb.add_format({"font_size": BRAND["body_pt"], "bold": True, "font_color": "#ffffff",
                                  "bg_color": deep, "align": "right", "indent": 1, "num_format": numfmt,
                                  "border": 1, "border_color": deep})
        f_tot_int, f_tot_days, f_tot_hrs = _tot("0"), _tot("0.0"), _tot('0.00" h"')
        f_tot_blank = wb.add_format({"bg_color": deep, "border": 1, "border_color": deep})
        for i, (_label, key, kind) in enumerate(COLS):
            if i == 0:
                ws.write(tr, i, "TOTAL", f_tot_l)
            elif kind in ("int", "days", "hours") and key != "avg_working_hours":
                colL = _xlc(i)
                tf = f_tot_hrs if kind == "hours" else (f_tot_days if kind == "days" else f_tot_int)
                ws.write_formula(tr, i, f"=SUM({colL}{start_row + 2}:{colL}{last_row + 1})", tf)
            else:
                ws.write_blank(tr, i, None, f_tot_blank)

    # ── Conditional heat-maps + data bars on the payroll-critical columns ──
    if rows:
        def cf(key, kw):
            ws.conditional_format(start_row + 1, col_idx[key], last_row, col_idx[key], kw)
        cf("late_days",   {"type": "3_color_scale", "min_color": "#ffffff", "mid_color": "#fef9c3", "max_color": "#f59e0b"})
        cf("absent_days", {"type": "3_color_scale", "min_color": "#ffffff", "mid_color": "#fecaca", "max_color": "#b91c1c"})
        cf("lop_days",    {"type": "3_color_scale", "min_color": "#ffffff", "mid_color": "#fecaca", "max_color": "#dc2626"})
        cf("excess_break_minutes", {"type": "3_color_scale", "min_color": "#ffffff", "mid_color": "#ffedd5", "max_color": "#ea580c"})
        cf("attendance_pct", {"type": "3_color_scale", "min_color": "#fecaca", "mid_color": "#fef9c3", "max_color": "#86efac"})
        cf("total_working_hours", {"type": "data_bar", "bar_color": accent, "bar_solid": False})
        cf("total_overtime_hours", {"type": "data_bar", "bar_color": "#ea580c", "bar_solid": True})
        ws.autofilter(start_row, 0, last_row, ncol - 1)

    # ─── Embedded charts ───
    if rows:
        name_c = col_idx["employee_name"]
        work_c = col_idx["total_working_hours"]
        ot_c = col_idx["total_overtime_hours"]
        lop_c = col_idx["lop_days"]

        chart = wb.add_chart({"type": "bar"})
        chart.add_series({
            "name": "Working hours",
            "categories": ["Monthly Summary", start_row + 1, name_c, last_row, name_c],
            "values":     ["Monthly Summary", start_row + 1, work_c, last_row, work_c],
            "fill": {"color": accent}, "border": {"color": deep},
        })
        chart.add_series({
            "name": "Overtime hours",
            "categories": ["Monthly Summary", start_row + 1, name_c, last_row, name_c],
            "values":     ["Monthly Summary", start_row + 1, ot_c, last_row, ot_c],
            "fill": {"color": "#ea580c"}, "border": {"color": "#7c2d12"},
        })
        chart.set_title({"name": "Working vs Overtime hours", "name_font": {"size": 13, "bold": True, "color": deep}})
        chart.set_x_axis({"name": "Hours", "num_format": "0"})
        chart.set_y_axis({"name": "Employee"})
        chart.set_legend({"position": "bottom"})
        chart.set_size({"width": 760, "height": max(360, 40 + 30 * len(rows))})

        lop_chart = wb.add_chart({"type": "column"})
        lop_chart.add_series({
            "name": "LOP days",
            "categories": ["Monthly Summary", start_row + 1, name_c, last_row, name_c],
            "values":     ["Monthly Summary", start_row + 1, lop_c, last_row, lop_c],
            "fill": {"color": "#dc2626"}, "border": {"color": "#7f1d1d"},
            "data_labels": {"value": True, "num_format": "0.0", "font": {"size": 8, "bold": True}},
        })
        lop_chart.set_title({"name": "Loss-of-Pay days by employee", "name_font": {"size": 13, "bold": True, "color": "#7f1d1d"}})
        lop_chart.set_legend({"none": True})
        lop_chart.set_size({"width": 760, "height": 320})

        chart_ws = wb.add_worksheet("Charts")
        chart_ws.set_tab_color("#fb923c")
        chart_ws.hide_gridlines(2)
        chart_ws.insert_chart("B2", chart)
        chart_ws.insert_chart("B22", lop_chart)

    return _xw_finalize(wb, buf)


# ════════════════════════════════════════════════════════════════════════════
# 2. Late Arrivals — xlsxwriter with HEAT-MAP scale
# ════════════════════════════════════════════════════════════════════════════


def _excel_late(rows: list[dict], summary: dict, meta: dict) -> bytes:
    theme = report_meta("late")
    accent = theme["accent"]
    deep = theme["accent_deep"]
    period = meta["period"]

    wb, buf = _xw_workbook()
    ws = wb.add_worksheet("Late Arrivals")
    ws.set_tab_color(accent)
    ws.hide_gridlines(2)

    ws.set_column("A:A", 13)
    ws.set_column("B:B", 12)
    ws.set_column("C:C", 26)
    ws.set_column("D:D", 22)
    ws.set_column("E:E", 18)
    ws.set_column("F:F", 13)
    ws.set_column("G:G", 12)
    ws.set_column("H:H", 14)

    # ── Corporate title block (rows 0..3) ──────────────────────────────
    next_row = _corporate_title_block(wb, ws, theme, period, summary, last_col=7)

    # ── KPI strip — punctuality vitals ─────────────────────────────────
    next_row = _corporate_kpi_strip(wb, ws, theme, [
        ("BREACHES",       summary["late"],         "#b91c1c"),
        ("LATE MINUTES",   summary["late_minutes"], "#d97706"),
        ("EMPLOYEES",      summary["employees"],    "#475569"),
        ("AVG LATE / EMP", round(summary["late_minutes"] / max(summary["employees"], 1)),
                                                    "#ea580c"),
    ], start_row=next_row, last_col=7)

    # ── Body formats ───────────────────────────────────────────────────
    f_header = _corporate_header_format(wb, theme, align="left")
    f_header_r = _corporate_header_format(wb, theme, align="right")
    f_cell = wb.add_format({"font_size": BRAND["body_pt"], "indent": 1, "border": 1, "border_color": BRAND["rule_soft"]})
    f_cell_z = wb.add_format({"font_size": BRAND["body_pt"], "indent": 1, "bg_color": BRAND["cream"], "border": 1, "border_color": BRAND["rule_soft"]})
    f_num = wb.add_format({"font_size": BRAND["body_pt"], "align": "right", "indent": 1, "num_format": "0", "border": 1, "border_color": BRAND["rule_soft"]})
    f_num_z = wb.add_format({"font_size": BRAND["body_pt"], "align": "right", "indent": 1, "num_format": "0", "bg_color": BRAND["cream"], "border": 1, "border_color": BRAND["rule_soft"]})
    f_date = wb.add_format({"font_size": BRAND["body_pt"], "indent": 1, "num_format": "dd mmm yyyy", "border": 1, "border_color": BRAND["rule_soft"]})
    f_date_z = wb.add_format({"font_size": BRAND["body_pt"], "indent": 1, "num_format": "dd mmm yyyy", "bg_color": BRAND["cream"], "border": 1, "border_color": BRAND["rule_soft"]})
    f_time = wb.add_format({"font_size": BRAND["body_pt"], "indent": 1, "num_format": "hh:mm AM/PM", "border": 1, "border_color": BRAND["rule_soft"]})
    f_time_z = wb.add_format({"font_size": BRAND["body_pt"], "indent": 1, "num_format": "hh:mm AM/PM", "bg_color": BRAND["cream"], "border": 1, "border_color": BRAND["rule_soft"]})

    # Headers
    headers = ["Date", "Code", "Employee", "Department", "Shift", "Check-in", "Late mins", "Status"]
    start_row = next_row
    ws.set_row(start_row, 24)
    for i, h in enumerate(headers):
        ws.write(start_row, i, h, f_header_r if h == "Late mins" else f_header)
    ws.freeze_panes(start_row + 1, 0)

    # Body
    for ri, r in enumerate(rows):
        row_idx = start_row + 1 + ri
        z = ri % 2 == 1
        ws.set_row(row_idx, 18)
        ws.write_datetime(row_idx, 0, _to_dt(r["date"]), f_date_z if z else f_date)
        ws.write(row_idx, 1, r["employee_code"], f_cell_z if z else f_cell)
        ws.write(row_idx, 2, r["employee_name"], f_cell_z if z else f_cell)
        ws.write(row_idx, 3, r["department"], f_cell_z if z else f_cell)
        ws.write(row_idx, 4, r["shift_name"], f_cell_z if z else f_cell)
        if r.get("check_in_time"):
            ws.write_datetime(row_idx, 5, r["check_in_time"], f_time_z if z else f_time)
        else:
            ws.write(row_idx, 5, "", f_cell_z if z else f_cell)
        ws.write_number(row_idx, 6, r["late_minutes"], f_num_z if z else f_num)
        _write_status_pill(wb, ws, row_idx, 7, r["status"], z)

    last_row = start_row + len(rows)
    if rows:
        # Heat-map on late minutes
        ws.conditional_format(start_row + 1, 6, last_row, 6, {
            "type": "3_color_scale",
            "min_type": "num", "min_value": 0, "min_color": "#ffffff",
            "mid_type": "num", "mid_value": 30, "mid_color": "#fef9c3",
            "max_type": "num", "max_value": 90, "max_color": "#b91c1c",
        })
        # Data bar
        ws.conditional_format(start_row + 1, 6, last_row, 6, {
            "type": "data_bar", "bar_color": "#ca8a04", "bar_only": False,
        })
        ws.autofilter(start_row, 0, last_row, len(headers) - 1)

    return _xw_finalize(wb, buf)


# ════════════════════════════════════════════════════════════════════════════
# 3. Overtime — xlsxwriter with embedded LINE CHART of OT per day
# ════════════════════════════════════════════════════════════════════════════


def _excel_overtime(rows: list[dict], summary: dict, meta: dict) -> bytes:
    theme = report_meta("overtime")
    accent = theme["accent"]
    deep = theme["accent_deep"]
    period = meta["period"]

    wb, buf = _xw_workbook()
    ws = wb.add_worksheet("Overtime")
    ws.set_tab_color(accent)
    ws.hide_gridlines(2)

    ws.set_column("A:A", 13)
    ws.set_column("B:B", 12)
    ws.set_column("C:C", 26)
    ws.set_column("D:D", 22)
    ws.set_column("E:E", 18)
    ws.set_column("F:F", 13)
    ws.set_column("G:H", 13)
    ws.set_column("I:I", 14)

    # ── Corporate title block ──────────────────────────────────────────
    next_row = _corporate_title_block(wb, ws, theme, period, summary, last_col=8)

    # ── KPI strip — operations vitals ──────────────────────────────────
    next_row = _corporate_kpi_strip(wb, ws, theme, [
        ("RECORDS",     summary["rows"],                       "#475569"),
        ("EMPLOYEES",   summary["employees"],                  "#0284c7"),
        ("ON-TIME %",   summary["on_time_pct"],                "#0d9488"),
        ("TOTAL OT",    round(summary["overtime_hours"], 1),   accent),
    ], start_row=next_row, last_col=8)

    # ── Body formats ───────────────────────────────────────────────────
    f_header = _corporate_header_format(wb, theme, align="left")
    f_header_r = _corporate_header_format(wb, theme, align="right")
    f_cell = wb.add_format({"font_size": BRAND["body_pt"], "indent": 1, "border": 1, "border_color": BRAND["rule_soft"]})
    f_cell_z = wb.add_format({"font_size": BRAND["body_pt"], "indent": 1, "bg_color": BRAND["cream"], "border": 1, "border_color": BRAND["rule_soft"]})
    f_hrs = wb.add_format({"font_size": BRAND["body_pt"], "align": "right", "indent": 1, "num_format": "0.00\" h\"", "bottom": 1, "bottom_color": BRAND["rule_soft"]})
    f_hrs_z = wb.add_format({"font_size": BRAND["body_pt"], "align": "right", "indent": 1, "num_format": "0.00\" h\"", "bg_color": BRAND["cream"], "bottom": 1, "bottom_color": BRAND["rule_soft"]})
    f_date = wb.add_format({"font_size": BRAND["body_pt"], "indent": 1, "num_format": "dd mmm yyyy", "bottom": 1, "bottom_color": BRAND["rule_soft"]})
    f_date_z = wb.add_format({"font_size": BRAND["body_pt"], "indent": 1, "num_format": "dd mmm yyyy", "bg_color": BRAND["cream"], "bottom": 1, "bottom_color": BRAND["rule_soft"]})
    f_time = wb.add_format({"font_size": BRAND["body_pt"], "indent": 1, "num_format": "hh:mm AM/PM", "bottom": 1, "bottom_color": BRAND["rule_soft"]})
    f_time_z = wb.add_format({"font_size": BRAND["body_pt"], "indent": 1, "num_format": "hh:mm AM/PM", "bg_color": BRAND["cream"], "bottom": 1, "bottom_color": BRAND["rule_soft"]})

    # Headers
    headers = ["Date", "Code", "Employee", "Department", "Shift",
               "Check-in", "Check-out", "OT hours", "Working hrs"]
    start_row = next_row
    ws.set_row(start_row, 24)
    for i, h in enumerate(headers):
        ws.write(start_row, i, h, f_header_r if h in ("OT hours", "Working hrs") else f_header)
    ws.freeze_panes(start_row + 1, 0)

    for ri, r in enumerate(rows):
        row_idx = start_row + 1 + ri
        z = ri % 2 == 1
        ws.set_row(row_idx, 18)
        ws.write_datetime(row_idx, 0, _to_dt(r["date"]), f_date_z if z else f_date)
        ws.write(row_idx, 1, r["employee_code"], f_cell_z if z else f_cell)
        ws.write(row_idx, 2, r["employee_name"], f_cell_z if z else f_cell)
        ws.write(row_idx, 3, r["department"], f_cell_z if z else f_cell)
        ws.write(row_idx, 4, r["shift_name"], f_cell_z if z else f_cell)
        if r.get("check_in_time"):
            ws.write_datetime(row_idx, 5, r["check_in_time"], f_time_z if z else f_time)
        else:
            ws.write(row_idx, 5, "", f_cell_z if z else f_cell)
        if r.get("check_out_time"):
            ws.write_datetime(row_idx, 6, r["check_out_time"], f_time_z if z else f_time)
        else:
            ws.write(row_idx, 6, "", f_cell_z if z else f_cell)
        ws.write_number(row_idx, 7, r["overtime_hours"], f_hrs_z if z else f_hrs)
        ws.write_number(row_idx, 8, r["working_hours"], f_hrs_z if z else f_hrs)

    last_row = start_row + len(rows)
    if rows:
        ws.conditional_format(start_row + 1, 7, last_row, 7, {
            "type": "data_bar", "bar_color": accent, "bar_solid": True,
        })
        ws.conditional_format(start_row + 1, 8, last_row, 8, {
            "type": "data_bar", "bar_color": "#fb923c", "bar_solid": False,
        })
        ws.autofilter(start_row, 0, last_row, len(headers) - 1)

        # Embedded column chart on a second sheet
        chart_ws = wb.add_worksheet("Chart · OT by employee")
        chart_ws.set_tab_color("#fb923c")
        chart_ws.hide_gridlines(2)

        chart = wb.add_chart({"type": "column"})
        chart.add_series({
            "name": "OT hours",
            "categories": ["Overtime", start_row + 1, 2, last_row, 2],
            "values":     ["Overtime", start_row + 1, 7, last_row, 7],
            "fill": {"color": accent},
            "border": {"color": deep},
            "data_labels": {"value": True, "num_format": "0.0", "font": {"size": 8}},
        })
        chart.set_title({"name": "Overtime hours per employee", "name_font": {"size": 13, "bold": True, "color": deep}})
        chart.set_x_axis({"name": "Employee"})
        chart.set_y_axis({"name": "OT hours", "num_format": "0.0"})
        chart.set_legend({"none": True})
        chart.set_size({"width": 840, "height": 480})
        chart_ws.insert_chart("B2", chart)

    return _xw_finalize(wb, buf)


# ════════════════════════════════════════════════════════════════════════════
# 4. WFH — openpyxl with calendar-feel merged header
# ════════════════════════════════════════════════════════════════════════════


def _excel_wfh(rows: list[dict], summary: dict, meta: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    theme = report_meta("wfh")
    accent = theme["accent"].lstrip("#")
    accent_soft = theme["accent_soft"].lstrip("#")
    deep = theme["accent_deep"].lstrip("#")
    period = meta["period"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Work From Home"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = accent

    thin = Side(style="thin", color="94A3B8")
    rule_soft = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell_border = Border(left=rule_soft, right=rule_soft, top=rule_soft, bottom=rule_soft)

    # ── Corporate title block (rows 1..4) ──────────────────────────────
    # Row 1 — 4pt accent rail
    ws.merge_cells("A1:H1")
    ws["A1"].fill = PatternFill("solid", fgColor=accent)
    ws.row_dimensions[1].height = 4

    # Row 2 — title (left) + report key (right)
    ws.merge_cells("A2:H2")
    ws["A2"] = f"  {COMPANY['name']}  ·  {theme['name']}"
    ws["A2"].font = Font(name="Calibri", size=18, bold=True, color="111418")
    ws["A2"].fill = PatternFill("solid", fgColor="FFFFFF")
    ws["A2"].alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws["A2"].border = Border(bottom=Side(style="thin", color="D1D5DB"))
    for col in "BCDEFGH":
        ws[f"{col}2"].border = Border(bottom=Side(style="thin", color="D1D5DB"))
        ws[f"{col}2"].fill = PatternFill("solid", fgColor="FFFFFF")
    ws.row_dimensions[2].height = 32

    # Row 3 — subtitle (italic)
    ws.merge_cells("A3:H3")
    ws["A3"] = f"  {theme['subtitle']}"
    ws["A3"].font = Font(name="Calibri", size=10, italic=True, color="475569")
    ws["A3"].fill = PatternFill("solid", fgColor="FFFFFF")
    ws["A3"].alignment = Alignment(vertical="center", horizontal="left", indent=1)
    for col in "BCDEFGH":
        ws[f"{col}3"].fill = PatternFill("solid", fgColor="FFFFFF")
    ws.row_dimensions[3].height = 20

    # Row 4 — period band
    ws.merge_cells("A4:H4")
    ws["A4"] = (
        f"  Period   {period['from'].strftime('%d %b %Y')}    →    {period['to'].strftime('%d %b %Y')}"
        f"        ·        Generated   {datetime.now().strftime('%d %b %Y · %I:%M %p').lstrip('0')}"
    )
    ws["A4"].font = Font(name="Calibri", size=9, color="475569")
    ws["A4"].fill = PatternFill("solid", fgColor="F8FAFC")
    ws["A4"].alignment = Alignment(vertical="center", horizontal="left", indent=1)
    for col in "BCDEFGH":
        ws[f"{col}4"].fill = PatternFill("solid", fgColor="F8FAFC")
    ws.row_dimensions[4].height = 22

    # ── KPI strip (rows 5 spacer · 6 labels · 7 values · 8 spacer) ─────
    ws.row_dimensions[5].height = 8
    rail_colors = ["0284C7", "0EA5E9", "475569", "0D9488", accent, "EA580C"]
    kpis = [
        ("WFH DAYS",     summary["wfh"]),
        ("REMOTE DAYS",  summary["remote"]),
        ("EMPLOYEES",    summary["employees"]),
        ("DEPARTMENTS",  summary["departments"]),
        ("WORKING HRS",  round(summary["working_hours"], 1)),
        ("OT HRS",       round(summary["overtime_hours"], 1)),
    ]
    tile_spans = [("A", "B"), ("C", "C"), ("D", "D"), ("E", "E"), ("F", "F"), ("G", "H")]
    for i, ((c0, c1), (lbl, val)) in enumerate(zip(tile_spans, kpis)):
        rail = rail_colors[i]
        # Label cell(s) — colored thick top rail + medium side rules
        if c0 != c1:
            ws.merge_cells(f"{c0}6:{c1}6")
        cell_l = ws[f"{c0}6"]
        cell_l.value = lbl
        cell_l.font = Font(name="Calibri", size=8, bold=True, color="475569")
        cell_l.fill = PatternFill("solid", fgColor="FFFFFF")
        cell_l.alignment = Alignment(vertical="center", horizontal="center")
        cell_l.border = Border(
            top=Side(style="thick", color=rail),
            left=Side(style="medium", color="94A3B8"),
            right=Side(style="medium", color="94A3B8"),
        )
        # Value cell(s) — medium side + bottom rule
        if c0 != c1:
            ws.merge_cells(f"{c0}7:{c1}7")
        cell_v = ws[f"{c0}7"]
        cell_v.value = val
        cell_v.font = Font(name="Calibri", size=18, bold=True, color="111418")
        cell_v.fill = PatternFill("solid", fgColor="FFFFFF")
        cell_v.alignment = Alignment(vertical="center", horizontal="center")
        cell_v.border = Border(
            left=Side(style="medium", color="94A3B8"),
            right=Side(style="medium", color="94A3B8"),
            bottom=Side(style="medium", color="94A3B8"),
        )
    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 30
    ws.row_dimensions[8].height = 10

    # Headers
    headers = ["Date", "Code", "Employee", "Department", "Check-in", "Check-out", "Working hrs", "Status"]
    head_row = 9
    deep_side = Side(style="medium", color=deep)
    for i, h in enumerate(headers):
        c = ws.cell(row=head_row, column=i + 1, value=h.upper())
        c.font = Font(size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=accent)
        c.alignment = Alignment(vertical="center",
                                horizontal="right" if h == "Working hrs" else "left",
                                indent=1)
        c.border = Border(top=deep_side, bottom=deep_side,
                          left=Side(style="thin", color=deep),
                          right=Side(style="thin", color=deep))
    ws.row_dimensions[head_row].height = 26
    ws.freeze_panes = f"A{head_row + 1}"

    # Body
    for ri, r in enumerate(rows):
        row_idx = head_row + 1 + ri
        is_zebra = ri % 2 == 1
        fill = PatternFill("solid", fgColor="FBF8F0") if is_zebra else PatternFill("solid", fgColor="FFFFFF")
        for col in range(1, 9):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = Font(size=10, color="111418")
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", indent=1)
            cell.border = cell_border
        ws.cell(row=row_idx, column=1, value=r["date"]).number_format = "dd mmm yyyy"
        ws.cell(row=row_idx, column=2, value=r["employee_code"])
        ws.cell(row=row_idx, column=3, value=r["employee_name"])
        ws.cell(row=row_idx, column=4, value=r["department"])
        if r.get("check_in_time"):
            ws.cell(row=row_idx, column=5, value=r["check_in_time"]).number_format = "hh:mm AM/PM"
        if r.get("check_out_time"):
            ws.cell(row=row_idx, column=6, value=r["check_out_time"]).number_format = "hh:mm AM/PM"
        hr_cell = ws.cell(row=row_idx, column=7, value=r["working_hours"])
        hr_cell.number_format = '0.00" h"'
        hr_cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        st = r["status"]
        sc = STATUS_COLORS.get(st, {"light": "#f1f5f9", "deep": "#334155"})
        scell = ws.cell(row=row_idx, column=8, value=st.replace("_", " "))
        scell.fill = PatternFill("solid", fgColor=sc["light"].lstrip("#"))
        scell.font = Font(size=10, bold=True, color=sc["deep"].lstrip("#"))
        scell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    widths = [14, 13, 28, 22, 13, 13, 14, 16]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # Autofilter
    if rows:
        ws.auto_filter.ref = f"A{head_row}:{get_column_letter(len(headers))}{head_row + len(rows)}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ════════════════════════════════════════════════════════════════════════════
# 5. Compliance — openpyxl with TRAFFIC-LIGHT coverage column
# ════════════════════════════════════════════════════════════════════════════


def _excel_compliance(rows: list[dict], summary: dict, meta: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

    theme = report_meta("compliance")
    accent = theme["accent"].lstrip("#")
    accent_soft = theme["accent_soft"].lstrip("#")
    deep = theme["accent_deep"].lstrip("#")
    period = meta["period"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compliance Audit"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = accent

    # ── Corporate title block ──────────────────────────────────────────
    ws.merge_cells("A1:I1")
    ws["A1"].fill = PatternFill("solid", fgColor=accent)
    ws.row_dimensions[1].height = 4

    ws.merge_cells("A2:I2")
    ws["A2"] = f"  {COMPANY['name']}  ·  {theme['name']}"
    ws["A2"].font = Font(name="Calibri", size=18, bold=True, color="111418")
    ws["A2"].fill = PatternFill("solid", fgColor="FFFFFF")
    ws["A2"].alignment = Alignment(vertical="center", horizontal="left", indent=1)
    for col in "BCDEFGHI":
        ws[f"{col}2"].fill = PatternFill("solid", fgColor="FFFFFF")
        ws[f"{col}2"].border = Border(bottom=Side(style="thin", color="D1D5DB"))
    ws["A2"].border = Border(bottom=Side(style="thin", color="D1D5DB"))
    ws.row_dimensions[2].height = 32

    ws.merge_cells("A3:I3")
    ws["A3"] = f"  {theme['subtitle']}"
    ws["A3"].font = Font(name="Calibri", size=10, italic=True, color="475569")
    ws["A3"].fill = PatternFill("solid", fgColor="FFFFFF")
    ws["A3"].alignment = Alignment(vertical="center", horizontal="left", indent=1)
    for col in "BCDEFGHI":
        ws[f"{col}3"].fill = PatternFill("solid", fgColor="FFFFFF")
    ws.row_dimensions[3].height = 20

    # Note: we keep the existing meta line below — it lands on row 4 already.
    ws.merge_cells("A4:I4")
    ws["A4"] = (
        f"  Period   {period['from'].strftime('%d %b %Y')}    →    {period['to'].strftime('%d %b %Y')}"
        f"        ·        Compiled   {datetime.now().strftime('%d %b %Y · %I:%M %p').lstrip('0')}"
        f"        ·        Ref FRC/HR/COMP/{period['from'].year}/{datetime.now().strftime('%m%d%H%M')}"
    )
    ws["A4"].font = Font(name="Calibri", size=9, color="475569")
    ws["A4"].fill = PatternFill("solid", fgColor="F8FAFC")
    ws["A4"].alignment = Alignment(vertical="center", horizontal="left", indent=1)
    for col in "BCDEFGHI":
        ws[f"{col}4"].fill = PatternFill("solid", fgColor="F8FAFC")
    ws.row_dimensions[4].height = 22

    # KPI ribbon — corporate tiles with colored top rails
    kpis = [
        ("EMPLOYEES",      summary["employees"],          "475569"),
        ("DEPARTMENTS",    summary["departments"],        "0284C7"),
        ("ON-TIME %",      f"{summary['on_time_pct']}%", "0D9488"),
        ("PRESENT",        summary["present"],           "0D9488"),
        ("LATE",           summary["late"],              "D97706"),
        ("ABSENT",         summary["absent"],            "B91C1C"),
    ]
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 30
    ws.row_dimensions[7].height = 12
    for i, (lbl, val, rail) in enumerate(kpis):
        col = i + 1 + ((9 - len(kpis)) // 2)  # center 6 tiles in 9 cols
        if col > 9:
            break
        # Label cell with colored thick top rail + medium side rules
        l_cell = ws.cell(row=5, column=col, value=lbl)
        l_cell.font = Font(name="Calibri", size=8, bold=True, color="475569")
        l_cell.fill = PatternFill("solid", fgColor="FFFFFF")
        l_cell.alignment = Alignment(vertical="center", horizontal="center")
        l_cell.border = Border(
            top=Side(style="thick", color=rail),
            left=Side(style="medium", color="94A3B8"),
            right=Side(style="medium", color="94A3B8"),
        )
        # Value cell — medium side + bottom rule for box look
        v_cell = ws.cell(row=6, column=col, value=val)
        v_cell.font = Font(name="Calibri", size=18, bold=True, color="111418")
        v_cell.fill = PatternFill("solid", fgColor="FFFFFF")
        v_cell.alignment = Alignment(vertical="center", horizontal="center")
        v_cell.border = Border(
            left=Side(style="medium", color="94A3B8"),
            right=Side(style="medium", color="94A3B8"),
            bottom=Side(style="medium", color="94A3B8"),
        )

    # Headers — corporate accent fill with dark side + bottom rules
    headers = [
        "Code", "Employee", "Department", "Shift",
        "Scheduled", "Actual hrs", "Expected hrs", "Coverage %", "Missing",
    ]
    head_row = 8
    deep_side = Side(style="medium", color=deep)
    for i, h in enumerate(headers):
        c = ws.cell(row=head_row, column=i + 1, value=h.upper())
        c.font = Font(size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=accent)
        c.alignment = Alignment(
            vertical="center",
            horizontal="right" if h in ("Scheduled", "Actual hrs", "Expected hrs", "Coverage %", "Missing") else "left",
            indent=1,
        )
        c.border = Border(top=deep_side, bottom=deep_side,
                          left=Side(style="thin", color=deep),
                          right=Side(style="thin", color=deep))
    ws.row_dimensions[head_row].height = 26
    ws.freeze_panes = f"A{head_row + 1}"

    # Body — strong grid borders so every cell reads as a discrete grid cell
    cell_grid = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    for ri, r in enumerate(rows):
        row_idx = head_row + 1 + ri
        zebra = ri % 2 == 1
        bg = "FBF8F0" if zebra else "FFFFFF"
        for col in range(1, 10):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = Font(size=10, color="111418")
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="center", indent=1)
            cell.border = cell_grid
        ws.cell(row=row_idx, column=1, value=r["employee_code"])
        ws.cell(row=row_idx, column=2, value=r["employee_name"])
        ws.cell(row=row_idx, column=3, value=r["department"])
        ws.cell(row=row_idx, column=4, value=r["shift_name"])
        ws.cell(row=row_idx, column=5, value=r["scheduled_days"]).alignment = Alignment(horizontal="right", indent=1)
        c_actual = ws.cell(row=row_idx, column=6, value=r["actual_hours"])
        c_actual.number_format = '0.00" h"'
        c_actual.alignment = Alignment(horizontal="right", indent=1)
        c_exp = ws.cell(row=row_idx, column=7, value=r["expected_hours"])
        c_exp.number_format = '0.00" h"'
        c_exp.alignment = Alignment(horizontal="right", indent=1)
        c_cov = ws.cell(row=row_idx, column=8, value=r["coverage_pct"] / 100.0)
        c_cov.number_format = "0%"
        c_cov.alignment = Alignment(horizontal="right", indent=1)
        c_cov.font = Font(size=10, bold=True)
        c_miss = ws.cell(row=row_idx, column=9, value=r["missing_punch_days"])
        c_miss.alignment = Alignment(horizontal="right", indent=1)
        if r["missing_punch_days"] > 0:
            c_miss.fill = PatternFill("solid", fgColor="FEE2E2")
            c_miss.font = Font(size=10, bold=True, color="7F1D1D")

    if rows:
        last = head_row + len(rows)
        # Traffic-light gradient on coverage column
        rule = ColorScaleRule(
            start_type="num", start_value=0.5, start_color="B91C1C",
            mid_type="num", mid_value=0.85, mid_color="FACC15",
            end_type="num", end_value=1.0, end_color="0D9488",
        )
        ws.conditional_formatting.add(f"H{head_row + 1}:H{last}", rule)
        ws.auto_filter.ref = f"A{head_row}:I{last}"

    widths = [12, 26, 22, 18, 12, 14, 14, 14, 11]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # Signature block
    sig_row = head_row + len(rows) + 4
    ws.row_dimensions[sig_row].height = 26
    ws.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row, end_column=4)
    ws.merge_cells(start_row=sig_row, start_column=6, end_row=sig_row, end_column=9)
    ws.cell(row=sig_row, column=1, value="").border = Border(top=Side(style="thin", color="1a1410"))
    ws.cell(row=sig_row + 1, column=1, value="Authorised by HR").alignment = Alignment(horizontal="center")
    ws.cell(row=sig_row + 1, column=1).font = Font(italic=True, color="4B5563", size=9, name="Georgia")
    ws.merge_cells(start_row=sig_row + 1, start_column=1, end_row=sig_row + 1, end_column=4)
    ws.cell(row=sig_row, column=6, value="").border = Border(top=Side(style="thin", color="1a1410"))
    ws.cell(row=sig_row + 1, column=6, value=f"Date · {datetime.now().strftime('%d %b %Y')}").alignment = Alignment(horizontal="center")
    ws.cell(row=sig_row + 1, column=6).font = Font(italic=True, color="4B5563", size=9, name="Georgia")
    ws.merge_cells(start_row=sig_row + 1, start_column=6, end_row=sig_row + 1, end_column=9)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ════════════════════════════════════════════════════════════════════════════
# 6. Anomalies — openpyxl with row-level severity tinting
# ════════════════════════════════════════════════════════════════════════════


def _excel_anomalies(rows: list[dict], summary: dict, meta: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    theme = report_meta("anomalies")
    accent = theme["accent"].lstrip("#")
    accent_soft = theme["accent_soft"].lstrip("#")
    deep = theme["accent_deep"].lstrip("#")
    period = meta["period"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Anomalies Dossier"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = accent

    # ── Corporate title block ──────────────────────────────────────────
    ws.merge_cells("A1:I1")
    ws["A1"].fill = PatternFill("solid", fgColor=accent)
    ws.row_dimensions[1].height = 4

    ws.merge_cells("A2:I2")
    ws["A2"] = f"  {COMPANY['name']}  ·  {theme['name']}"
    ws["A2"].font = Font(name="Calibri", size=18, bold=True, color="111418")
    ws["A2"].fill = PatternFill("solid", fgColor="FFFFFF")
    ws["A2"].alignment = Alignment(vertical="center", horizontal="left", indent=1)
    for col in "BCDEFGHI":
        ws[f"{col}2"].fill = PatternFill("solid", fgColor="FFFFFF")
        ws[f"{col}2"].border = Border(bottom=Side(style="thin", color="D1D5DB"))
    ws["A2"].border = Border(bottom=Side(style="thin", color="D1D5DB"))
    ws.row_dimensions[2].height = 32

    ws.merge_cells("A3:I3")
    ws["A3"] = f"  {theme['subtitle']}"
    ws["A3"].font = Font(name="Calibri", size=10, italic=True, color="475569")
    ws["A3"].fill = PatternFill("solid", fgColor="FFFFFF")
    ws["A3"].alignment = Alignment(vertical="center", horizontal="left", indent=1)
    for col in "BCDEFGHI":
        ws[f"{col}3"].fill = PatternFill("solid", fgColor="FFFFFF")
    ws.row_dimensions[3].height = 20

    ws.merge_cells("A4:I4")
    severity = "HIGH" if summary["late"] > 5 else "MODERATE" if summary["late"] > 0 else "LOW"
    sev_color = {"HIGH": "B91C1C", "MODERATE": "D97706", "LOW": "0D9488"}[severity]
    ws["A4"] = (
        f"  Case  FRC/HR/ANM/{period['from'].year}/{datetime.now().strftime('%m%d%H%M')}"
        f"        ·        Window  {period['from'].strftime('%d %b %Y')} → {period['to'].strftime('%d %b %Y')}"
        f"        ·        Severity  {severity}"
    )
    ws["A4"].font = Font(name="Calibri", size=9, color="475569", bold=False)
    ws["A4"].fill = PatternFill("solid", fgColor="F8FAFC")
    ws["A4"].alignment = Alignment(vertical="center", horizontal="left", indent=1)
    for col in "BCDEFGHI":
        ws[f"{col}4"].fill = PatternFill("solid", fgColor="F8FAFC")
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 14  # spacer before headers

    # Headers — corporate accent fill, dark borders on every side
    headers = ["Date", "Code", "Employee", "Department", "Status",
               "Check-in", "Check-out", "Late mins", "Reasons"]
    head_row = 6
    deep_side = Side(style="medium", color=deep)
    for i, h in enumerate(headers):
        c = ws.cell(row=head_row, column=i + 1, value=h.upper())
        c.font = Font(size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=accent)
        c.alignment = Alignment(vertical="center",
                                horizontal="right" if h == "Late mins" else "left",
                                indent=1)
        c.border = Border(top=deep_side, bottom=deep_side,
                          left=Side(style="thin", color=deep),
                          right=Side(style="thin", color=deep))
    ws.row_dimensions[head_row].height = 26
    ws.freeze_panes = f"A{head_row + 1}"

    # Body — manila row with severity stripe on left
    for ri, r in enumerate(rows):
        row_idx = head_row + 1 + ri
        zebra = ri % 2 == 1
        bg = "FAF4DC" if zebra else "FEF9EC"  # warmer manila
        # Severity strip color
        sev = r.get("severity", 0)
        stripe = accent if sev >= 5 else "B45309" if sev >= 3 else "92400E"

        cell_grid = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )
        for col in range(1, 10):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = Font(size=10, color="111418")
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="center", indent=1)
            cell.border = cell_grid

        # Severity stripe on date column — keep the thick left rail signal
        date_cell = ws.cell(row=row_idx, column=1, value=r["date"])
        date_cell.number_format = "dd mmm yyyy"
        date_cell.border = Border(
            left=Side(style="thick", color=stripe),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )
        date_cell.font = Font(size=10, bold=True, color="1A1410")
        ws.cell(row=row_idx, column=2, value=r["employee_code"])
        ws.cell(row=row_idx, column=3, value=r["employee_name"])
        ws.cell(row=row_idx, column=4, value=r["department"])

        st = r["status"]
        sc = STATUS_COLORS.get(st, {"light": "#f1f5f9", "deep": "#334155"})
        scell = ws.cell(row=row_idx, column=5, value=st.replace("_", " "))
        scell.fill = PatternFill("solid", fgColor=sc["light"].lstrip("#"))
        scell.font = Font(size=9, bold=True, color=sc["deep"].lstrip("#"))
        scell.alignment = Alignment(horizontal="center", vertical="center")

        if r.get("check_in_time"):
            ws.cell(row=row_idx, column=6, value=r["check_in_time"]).number_format = "hh:mm AM/PM"
        if r.get("check_out_time"):
            cout = ws.cell(row=row_idx, column=7, value=r["check_out_time"])
            cout.number_format = "hh:mm AM/PM"
        else:
            # Highlight missing check-out
            no_out = ws.cell(row=row_idx, column=7, value="—")
            no_out.fill = PatternFill("solid", fgColor="FECACA")
            no_out.font = Font(size=10, bold=True, color="7F1D1D")
            no_out.alignment = Alignment(horizontal="center", vertical="center")

        c_late = ws.cell(row=row_idx, column=8, value=r["late_minutes"])
        c_late.alignment = Alignment(horizontal="right", indent=1)
        if r["late_minutes"] > 30:
            c_late.fill = PatternFill("solid", fgColor="FEE2E2")
            c_late.font = Font(size=10, bold=True, color="7F1D1D")
        c_late.number_format = "0"

        rcell = ws.cell(row=row_idx, column=9, value=r.get("reasons", ""))
        rcell.font = Font(size=10, italic=True, color=deep)
        rcell.alignment = Alignment(vertical="center", indent=1, wrap_text=True)

    if rows:
        ws.auto_filter.ref = f"A{head_row}:I{head_row + len(rows)}"

    widths = [14, 12, 26, 22, 14, 13, 13, 11, 36]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ════════════════════════════════════════════════════════════════════════════
# 7. Daily Roster — xlsxwriter with autofilter + frozen panes + status pills
# ════════════════════════════════════════════════════════════════════════════


def _excel_daily(rows: list[dict], summary: dict, meta: dict) -> bytes:
    theme = report_meta("daily")
    accent = theme["accent"]
    deep = theme["accent_deep"]
    period = meta["period"]

    wb, buf = _xw_workbook()
    ws = wb.add_worksheet("Daily Roster")
    ws.set_tab_color(accent)
    ws.hide_gridlines(2)

    ws.set_column("A:A", 14)
    ws.set_column("B:B", 12)
    ws.set_column("C:C", 26)
    ws.set_column("D:D", 20)
    ws.set_column("E:E", 16)
    ws.set_column("F:G", 13)
    ws.set_column("H:K", 11)
    ws.set_column("L:L", 14)

    # ── Corporate title block ──────────────────────────────────────────
    next_row = _corporate_title_block(wb, ws, theme, period, summary, last_col=11)

    # ── KPI strip — daily roster vitals ────────────────────────────────
    next_row = _corporate_kpi_strip(wb, ws, theme, [
        ("EMPLOYEES",     summary["employees"],                   "#475569"),
        ("ROWS",          summary["rows"],                        "#0284c7"),
        ("ON-TIME %",     summary["on_time_pct"],                 "#0d9488"),
        ("LATE EVENTS",   summary["late"],                        "#d97706"),
        ("ABSENT",        summary["absent"],                      "#b91c1c"),
        ("OT HRS",        round(summary["overtime_hours"], 1),    "#ea580c"),
    ], start_row=next_row, last_col=11)

    # ── Body formats ───────────────────────────────────────────────────
    f_header = _corporate_header_format(wb, theme, align="left")
    f_header_r = _corporate_header_format(wb, theme, align="right")
    f_cell = wb.add_format({"font_size": BRAND["body_pt"], "indent": 1, "border": 1, "border_color": BRAND["rule_soft"]})
    f_cell_z = wb.add_format({"font_size": BRAND["body_pt"], "indent": 1, "bg_color": BRAND["cream"], "border": 1, "border_color": BRAND["rule_soft"]})
    f_hrs = wb.add_format({"font_size": BRAND["body_pt"], "align": "right", "indent": 1, "num_format": "0.00\" h\"", "bottom": 1, "bottom_color": BRAND["rule_soft"]})
    f_hrs_z = wb.add_format({"font_size": BRAND["body_pt"], "align": "right", "indent": 1, "num_format": "0.00\" h\"", "bg_color": BRAND["cream"], "bottom": 1, "bottom_color": BRAND["rule_soft"]})
    f_num = wb.add_format({"font_size": BRAND["body_pt"], "align": "right", "indent": 1, "num_format": "0", "border": 1, "border_color": BRAND["rule_soft"]})
    f_num_z = wb.add_format({"font_size": BRAND["body_pt"], "align": "right", "indent": 1, "num_format": "0", "bg_color": BRAND["cream"], "border": 1, "border_color": BRAND["rule_soft"]})
    f_date = wb.add_format({"font_size": BRAND["body_pt"], "indent": 1, "num_format": "dd mmm yyyy", "border": 1, "border_color": BRAND["rule_soft"]})
    f_date_z = wb.add_format({"font_size": BRAND["body_pt"], "indent": 1, "num_format": "dd mmm yyyy", "bg_color": BRAND["cream"], "border": 1, "border_color": BRAND["rule_soft"]})
    f_time = wb.add_format({"font_size": BRAND["body_pt"], "indent": 1, "num_format": "hh:mm AM/PM", "border": 1, "border_color": BRAND["rule_soft"]})
    f_time_z = wb.add_format({"font_size": BRAND["body_pt"], "indent": 1, "num_format": "hh:mm AM/PM", "bg_color": BRAND["cream"], "border": 1, "border_color": BRAND["rule_soft"]})

    # Headers
    headers = ["Date", "Code", "Employee", "Department", "Shift",
               "Check-in", "Check-out", "Hours", "Break", "Late", "OT", "Status"]
    start_row = next_row
    ws.set_row(start_row, 26)
    for i, h in enumerate(headers):
        ws.write(start_row, i, h, f_header_r if h in ("Hours", "Break", "Late", "OT") else f_header)
    ws.freeze_panes(start_row + 1, 0)

    # Body
    for ri, r in enumerate(rows):
        row_idx = start_row + 1 + ri
        z = ri % 2 == 1
        ws.set_row(row_idx, 18)
        ws.write_datetime(row_idx, 0, _to_dt(r["date"]), f_date_z if z else f_date)
        ws.write(row_idx, 1, r["employee_code"], f_cell_z if z else f_cell)
        ws.write(row_idx, 2, r["employee_name"], f_cell_z if z else f_cell)
        ws.write(row_idx, 3, r["department"], f_cell_z if z else f_cell)
        ws.write(row_idx, 4, r["shift_name"], f_cell_z if z else f_cell)
        if r.get("check_in_time"):
            ws.write_datetime(row_idx, 5, r["check_in_time"], f_time_z if z else f_time)
        else:
            ws.write(row_idx, 5, "", f_cell_z if z else f_cell)
        if r.get("check_out_time"):
            ws.write_datetime(row_idx, 6, r["check_out_time"], f_time_z if z else f_time)
        else:
            ws.write(row_idx, 6, "", f_cell_z if z else f_cell)
        ws.write_number(row_idx, 7, r["working_hours"], f_hrs_z if z else f_hrs)
        ws.write_number(row_idx, 8, r["break_hours"], f_hrs_z if z else f_hrs)
        ws.write_number(row_idx, 9, r["late_minutes"], f_num_z if z else f_num)
        ws.write_number(row_idx, 10, r["overtime_hours"], f_hrs_z if z else f_hrs)
        _write_status_pill(wb, ws, row_idx, 11, r["status"], z)

    last_row = start_row + len(rows)
    if rows:
        ws.conditional_format(start_row + 1, 8, last_row, 8, {
            "type": "data_bar", "bar_color": "#0284c7", "bar_solid": False,
        })
        ws.conditional_format(start_row + 1, 9, last_row, 9, {
            "type": "3_color_scale",
            "min_color": "#ffffff", "mid_color": "#fef9c3", "max_color": "#b91c1c",
        })
        ws.conditional_format(start_row + 1, 10, last_row, 10, {
            "type": "data_bar", "bar_color": "#fb923c",
        })
        ws.autofilter(start_row, 0, last_row, len(headers) - 1)

    return _xw_finalize(wb, buf)


# ════════════════════════════════════════════════════════════════════════════
# Shared helpers
# ════════════════════════════════════════════════════════════════════════════


def _to_dt(d):
    """Convert date/datetime/string to datetime (xlsxwriter requires datetime)."""
    if isinstance(d, datetime):
        return d
    if isinstance(d, date_cls):
        return datetime(d.year, d.month, d.day)
    if isinstance(d, str):
        try:
            return datetime.fromisoformat(d)
        except ValueError:
            return None
    return None


def _write_status_pill(wb, ws, row, col, status, zebra=False):
    """Write a status string with a pill-style background tinted by STATUS_COLORS."""
    sc = STATUS_COLORS.get(status, {"light": "#f1f5f9", "deep": "#334155"})
    fmt = wb.add_format({
        "font_size": 9, "bold": True, "italic": False,
        "font_color": sc["deep"],
        "bg_color": sc["light"],
        "align": "center", "valign": "vcenter",
        "bottom": 1, "bottom_color": "#ece6d7",
    })
    ws.write(row, col, str(status).replace("_", " "), fmt)


def _excel_breaks(rows: list[dict], summary: dict, meta: dict) -> bytes:
    """Breaks — cafe receipt style. Monospaced title band, dashed dividers,
    intensity tags in pill format, ratio data bar."""
    theme = report_meta("breaks")
    accent = theme["accent"]
    deep = theme["accent_deep"]
    period = meta["period"]

    wb, buf = _xw_workbook()
    ws = wb.add_worksheet("Breaks")
    ws.set_tab_color(accent)
    ws.hide_gridlines(2)

    ws.set_column("A:A", 14)
    ws.set_column("B:B", 12)
    ws.set_column("C:C", 26)
    ws.set_column("D:D", 22)
    ws.set_column("E:E", 18)
    ws.set_column("F:G", 14)
    ws.set_column("H:I", 12)
    ws.set_column("J:J", 14)

    # ── Corporate title block ──────────────────────────────────────────
    next_row = _corporate_title_block(wb, ws, theme, period, summary, last_col=9)

    # ── KPI strip — break intelligence ─────────────────────────────────
    total_break_min = sum(int(r.get("break_minutes") or 0) for r in rows)
    avg_ratio = round(sum(float(r.get("break_ratio_pct") or 0) for r in rows) / max(len(rows), 1))
    next_row = _corporate_kpi_strip(wb, ws, theme, [
        ("BREAK-DAYS",     len(rows),                "#475569"),
        ("EMPLOYEES",      summary["employees"],     "#0284c7"),
        ("TOTAL BREAK MIN", total_break_min,         accent),
        ("AVG RATIO %",    avg_ratio,                "#d97706"),
    ], start_row=next_row, last_col=9)

    # ── Body formats ───────────────────────────────────────────────────
    f_header = _corporate_header_format(wb, theme, align="left")
    f_header_r = _corporate_header_format(wb, theme, align="right")
    f_cell = wb.add_format({"font_size": BRAND["body_pt"], "indent": 1, "border": 1, "border_color": BRAND["rule_soft"]})
    f_cell_z = wb.add_format({"font_size": BRAND["body_pt"], "indent": 1, "bg_color": BRAND["cream"], "border": 1, "border_color": BRAND["rule_soft"]})
    f_hrs = wb.add_format({"font_size": BRAND["body_pt"], "align": "right", "indent": 1, "num_format": "0.00\" h\"", "bottom": 1, "bottom_color": BRAND["rule_soft"]})
    f_hrs_z = wb.add_format({"font_size": BRAND["body_pt"], "align": "right", "indent": 1, "num_format": "0.00\" h\"", "bg_color": BRAND["cream"], "bottom": 1, "bottom_color": BRAND["rule_soft"]})
    f_num = wb.add_format({"font_size": BRAND["body_pt"], "align": "right", "indent": 1, "num_format": "0", "bottom": 1, "bottom_color": BRAND["rule_soft"]})
    f_num_z = wb.add_format({"font_size": BRAND["body_pt"], "align": "right", "indent": 1, "num_format": "0", "bg_color": BRAND["cream"], "bottom": 1, "bottom_color": BRAND["rule_soft"]})
    f_pct = wb.add_format({"font_size": BRAND["body_pt"], "align": "right", "indent": 1, "num_format": "0\"%\"", "border": 1, "border_color": BRAND["rule_soft"]})
    f_pct_z = wb.add_format({"font_size": BRAND["body_pt"], "align": "right", "indent": 1, "num_format": "0\"%\"", "bg_color": BRAND["cream"], "border": 1, "border_color": BRAND["rule_soft"]})
    f_date = wb.add_format({"font_size": BRAND["body_pt"], "indent": 1, "num_format": "dd mmm yyyy", "bottom": 1, "bottom_color": BRAND["rule_soft"]})
    f_date_z = wb.add_format({"font_size": BRAND["body_pt"], "indent": 1, "num_format": "dd mmm yyyy", "bg_color": BRAND["cream"], "bottom": 1, "bottom_color": BRAND["rule_soft"]})

    # Intensity pill formats
    INTENSITY_STYLES = {
        "SHORT":    {"bg": BRAND["good_bg"], "fg": BRAND["good_fg"]},
        "STANDARD": {"bg": BRAND["warn_bg"], "fg": BRAND["warn_fg"]},
        "LONG":     {"bg": BRAND["danger_bg"], "fg": BRAND["danger_fg"]},
    }

    # Headers
    headers = ["Date", "Code", "Employee", "Department", "Shift",
               "Working hrs", "Break hrs", "Break mins", "Ratio %", "Length"]
    start_row = next_row
    ws.set_row(start_row, 26)
    for i, h in enumerate(headers):
        ws.write(start_row, i, h, f_header_r if h in ("Working hrs", "Break hrs", "Break mins", "Ratio %") else f_header)
    ws.freeze_panes(start_row + 1, 0)

    for ri, r in enumerate(rows):
        row_idx = start_row + 1 + ri
        z = ri % 2 == 1
        ws.set_row(row_idx, 18)
        ws.write_datetime(row_idx, 0, _to_dt(r["date"]), f_date_z if z else f_date)
        ws.write(row_idx, 1, r["employee_code"], f_cell_z if z else f_cell)
        ws.write(row_idx, 2, r["employee_name"], f_cell_z if z else f_cell)
        ws.write(row_idx, 3, r["department"], f_cell_z if z else f_cell)
        ws.write(row_idx, 4, r["shift_name"], f_cell_z if z else f_cell)
        ws.write_number(row_idx, 5, r["working_hours"], f_hrs_z if z else f_hrs)
        ws.write_number(row_idx, 6, r["break_hours"], f_hrs_z if z else f_hrs)
        ws.write_number(row_idx, 7, r["break_minutes"], f_num_z if z else f_num)
        ws.write_number(row_idx, 8, r["break_ratio_pct"], f_pct_z if z else f_pct)
        # Intensity pill
        st = INTENSITY_STYLES.get(r["intensity"], {"bg": "#f1f5f9", "fg": "#334155"})
        f_pill = wb.add_format({
            "font_size": 9, "bold": True, "italic": False,
            "font_color": st["fg"], "bg_color": st["bg"],
            "align": "center", "valign": "vcenter",
            "bottom": 1, "bottom_color": "#ece6d7",
        })
        ws.write(row_idx, 9, r["intensity"], f_pill)

    last_row = start_row + len(rows)
    if rows:
        # Break hours data bar (cafe brown)
        ws.conditional_format(start_row + 1, 6, last_row, 6, {
            "type": "data_bar", "bar_color": accent, "bar_solid": True,
        })
        # Ratio % heat-map
        ws.conditional_format(start_row + 1, 8, last_row, 8, {
            "type": "3_color_scale",
            "min_type": "num", "min_value": 0, "min_color": "#dcfce7",
            "mid_type": "num", "mid_value": 15, "mid_color": "#fef9c3",
            "max_type": "num", "max_value": 30, "max_color": "#fecaca",
        })
        ws.autofilter(start_row, 0, last_row, len(headers) - 1)

    return _xw_finalize(wb, buf)


RENDERERS = {
    "monthly": _excel_monthly,
    "late": _excel_late,
    "overtime": _excel_overtime,
    "wfh": _excel_wfh,
    "compliance": _excel_compliance,
    "anomalies": _excel_anomalies,
    "daily": _excel_daily,
    "breaks": _excel_breaks,
}

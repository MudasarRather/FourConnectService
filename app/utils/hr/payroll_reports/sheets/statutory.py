"""Excel workbook for the Statutory Summary report (openpyxl).

A bespoke compliance worksheet:
  · accent title rail + serif-feel title block + period line
  · merged-cell KPI tiles (PF / ESI / PT / TDS / Total) with colored top rails
  · themed header row with frozen panes + autofilter
  · bordered zebra body rows, money columns Indian-grouped
  · grouped sub-headers over the PF / ESI pairs
  · a TOTAL row summing every money column
  · a traffic-light ColorScaleRule heat-map on the Statutory total column
  · a CellIsRule highlight on the TDS column for high-deduction rows

openpyxl is chosen (per the package brief) for formula-friendly, edit-ready
formatting and the colour-scale conditional format.
"""
from __future__ import annotations

import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter

from ..data import report_meta


# ── theme accents (hex WITHOUT '#') ─────────────────────────────────────────
def _hx(c: str) -> str:
    return c.lstrip("#")


MONEY_FMT = '"Rs"#,##0;[Red]-"Rs"#,##0'
INK = "1f2933"
WHITE = "FFFFFF"
CREAM = "FBF7EE"
RULE_SOFT = "D6CDBA"
BRONZE = "9A6B2F"
BRONZE_DEEP = "6B4A1F"
DANGER = "B91C1C"


# Column layout: (header, row-key, kind)  kind in {text, money}
_COLS = [
    ("Emp Code",   "employee_code",  "text"),
    ("Employee",   "employee_name",  "text"),
    ("PAN",        "pan",            "text"),
    ("UAN",        "uan",            "text"),
    ("PF — EE",    "pf_employee",    "money"),
    ("PF — ER",    "pf_employer",    "money"),
    ("ESI — EE",   "esi_employee",   "money"),
    ("ESI — ER",   "esi_employer",   "money"),
    ("Prof. Tax",  "pt",             "money"),
    ("TDS",        "tds",            "money"),
    ("Statutory Total", "statutory_total", "money"),
]
_WIDTHS = [12, 26, 14, 16, 13, 13, 13, 13, 12, 13, 17]


def _thin(color: str = RULE_SOFT) -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def render(rows: list[dict], summary: dict, meta: dict) -> bytes:
    theme = report_meta("statutory")
    accent = _hx(theme.get("accent", "#0d9488"))
    deep = _hx(theme.get("accent_deep", "#134e4a"))
    soft = _hx(theme.get("accent_soft", "#ccfbf1"))
    name = theme.get("name", "Statutory Summary")
    subtitle = theme.get("subtitle", "")
    period = meta.get("period", {})
    plabel = period.get("label", "")
    fy = period.get("fy", "")

    ncols = len(_COLS)
    last_col = ncols  # 1-based count == last column index

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name[:31]
    ws.sheet_view.showGridLines = False
    if ws.sheet_properties.tabColor is not None or True:
        ws.sheet_properties.tabColor = accent

    for i, w in enumerate(_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Row 1: accent rail ───────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor=accent)
    ws.row_dimensions[1].height = 6

    # ── Row 2: title ─────────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    c = ws.cell(row=2, column=1, value=f"  Fourreck  ·  {name}")
    c.font = Font(name="Georgia", bold=True, size=18, color=INK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 30

    # ── Row 3: subtitle ──────────────────────────────────────────────────
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
    c = ws.cell(row=3, column=1, value=f"  {subtitle}")
    c.font = Font(italic=True, size=10.5, color="5B5345")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 18

    # ── Row 4: period line ───────────────────────────────────────────────
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=last_col)
    c = ws.cell(row=4, column=1,
                value=f"  Pay period   {plabel}      ·      FY {fy}      ·      Certified for statutory filing")
    c.font = Font(bold=True, size=9, color=deep)
    c.fill = PatternFill("solid", fgColor=soft)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = Border(bottom=Side(style="medium", color=deep))
    ws.row_dimensions[4].height = 20

    # ── KPI tiles (rows 6-7), merged with colored top rail ──────────────
    pf = float(summary.get("pf", 0) or 0)
    esi = float(summary.get("esi", 0) or 0)
    pt = float(summary.get("pt", 0) or 0)
    tds = float(summary.get("tds", 0) or 0)
    emp = summary.get("employees", summary.get("rows", len(rows)))
    total = pf + esi + pt + tds

    # (label, value, is_money, rail_hex)
    kpis = [
        ("EMPLOYEES", emp, False, deep),
        ("PF REMITTED", pf, True, accent),
        ("ESI REMITTED", esi, True, accent),
        ("PROF. TAX", pt, True, BRONZE),
        ("TDS DEDUCTED", tds, True, DANGER),
        ("TOTAL STATUTORY", total, True, deep),
    ]
    klabel_row, kval_row = 6, 7
    ws.row_dimensions[5].height = 6
    ws.row_dimensions[klabel_row].height = 16
    ws.row_dimensions[kval_row].height = 26

    n = len(kpis)
    per = max(1, ncols // n)
    leftover = ncols - per * n
    col = 1
    for idx, (label, value, is_money, rail) in enumerate(kpis):
        span = per + (1 if idx < leftover else 0)
        c0, c1 = col, col + span - 1
        # label
        if c0 != c1:
            ws.merge_cells(start_row=klabel_row, start_column=c0, end_row=klabel_row, end_column=c1)
        lc = ws.cell(row=klabel_row, column=c0, value=label)
        lc.font = Font(bold=True, size=8, color="475569")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.fill = PatternFill("solid", fgColor=WHITE)
        for cc in range(c0, c1 + 1):
            ws.cell(row=klabel_row, column=cc).border = Border(top=Side(style="thick", color=rail))
        # value
        if c0 != c1:
            ws.merge_cells(start_row=kval_row, start_column=c0, end_row=kval_row, end_column=c1)
        vc = ws.cell(row=kval_row, column=c0, value=value)
        vc.font = Font(bold=True, size=15, color=INK)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.fill = PatternFill("solid", fgColor=WHITE)
        if is_money:
            vc.number_format = MONEY_FMT
        vc.border = Border(bottom=Side(style="thin", color=RULE_SOFT))
        col = c1 + 1

    # ── Grouped sub-headers (row 9) over PF / ESI pairs ─────────────────
    grp_row = 9
    ws.row_dimensions[8].height = 6
    grp_fill = PatternFill("solid", fgColor=deep)
    grp_font = Font(bold=True, size=8.5, color=WHITE)
    grp_align = Alignment(horizontal="center", vertical="center")

    def _group(c0: int, c1: int, text: str):
        ws.merge_cells(start_row=grp_row, start_column=c0, end_row=grp_row, end_column=c1)
        cell = ws.cell(row=grp_row, column=c0, value=text)
        cell.fill = grp_fill
        cell.font = grp_font
        cell.alignment = grp_align
        for cc in range(c0, c1 + 1):
            ws.cell(row=grp_row, column=cc).border = _thin(deep)

    # identity columns 1-4 -> one group
    _group(1, 4, "IDENTITY")
    _group(5, 6, "PROVIDENT FUND (EPF)")
    _group(7, 8, "EMPLOYEES' STATE INS. (ESI)")
    _group(9, 9, "PT")
    _group(10, 10, "TDS")
    _group(11, 11, "TOTAL")
    ws.row_dimensions[grp_row].height = 18

    # ── Header row (row 10) ──────────────────────────────────────────────
    head_row = grp_row + 1
    head_fill = PatternFill("solid", fgColor=accent)
    head_font = Font(bold=True, size=10, color=WHITE)
    for i, (label, _key, kind) in enumerate(_COLS, start=1):
        hc = ws.cell(row=head_row, column=i, value=label)
        hc.fill = head_fill
        hc.font = head_font
        hc.alignment = Alignment(horizontal="right" if kind == "money" else "left",
                                 vertical="center", indent=1)
        hc.border = Border(top=Side(style="medium", color=deep),
                           bottom=Side(style="medium", color=deep),
                           left=Side(style="thin", color=deep),
                           right=Side(style="thin", color=deep))
    ws.row_dimensions[head_row].height = 24
    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)

    # ── Body rows ────────────────────────────────────────────────────────
    first_data = head_row + 1
    for ri, r in enumerate(rows):
        rr = first_data + ri
        zebra = ri % 2 == 1
        fill = PatternFill("solid", fgColor=CREAM) if zebra else None
        for i, (label, key, kind) in enumerate(_COLS, start=1):
            v = r.get(key)
            if kind == "money":
                cell = ws.cell(row=rr, column=i, value=float(v or 0))
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right", indent=1, vertical="center")
            else:
                cell = ws.cell(row=rr, column=i, value=(v if v not in (None, "") else "—"))
                cell.alignment = Alignment(horizontal="left", indent=1, vertical="center")
            cell.font = Font(size=10, color=INK)
            cell.border = _thin()
            if fill is not None:
                cell.fill = fill
        ws.row_dimensions[rr].height = 16

    last_data = first_data + len(rows) - 1 if rows else head_row

    # ── TOTAL row ────────────────────────────────────────────────────────
    if rows:
        tr = last_data + 1
        tot_fill = PatternFill("solid", fgColor=deep)
        tot_font = Font(bold=True, size=10, color=WHITE)
        for i, (label, key, kind) in enumerate(_COLS, start=1):
            cell = ws.cell(row=tr, column=i)
            cell.fill = tot_fill
            cell.border = _thin(deep)
            if i == 1:
                cell.value = "TOTAL"
                cell.font = tot_font
                cell.alignment = Alignment(horizontal="left", indent=1, vertical="center")
            elif kind == "money":
                cl = get_column_letter(i)
                cell.value = f"=SUM({cl}{first_data}:{cl}{last_data})"
                cell.number_format = MONEY_FMT
                cell.font = tot_font
                cell.alignment = Alignment(horizontal="right", indent=1, vertical="center")
            else:
                cell.font = tot_font
        ws.row_dimensions[tr].height = 20

        # ── Conditional formats ──────────────────────────────────────────
        stat_col = get_column_letter(ncols)          # Statutory Total
        tds_col = get_column_letter(ncols - 1)       # TDS
        data_rng = f"{stat_col}{first_data}:{stat_col}{last_data}"
        # Traffic-light colour scale on statutory total (green→amber→red)
        ws.conditional_formatting.add(
            data_rng,
            ColorScaleRule(
                start_type="min", start_color="C9F2E9",
                mid_type="percentile", mid_value=50, mid_color="FDE9C8",
                end_type="max", end_color="F6C6C6",
            ),
        )
        # Highlight high TDS (> average proxy: top quartile cell flag)
        tds_rng = f"{tds_col}{first_data}:{tds_col}{last_data}"
        tds_vals = [float(r.get("tds") or 0) for r in rows]
        if tds_vals:
            thresh = sum(tds_vals) / len(tds_vals)
            ws.conditional_formatting.add(
                tds_rng,
                CellIsRule(operator="greaterThan", formula=[str(thresh)],
                           fill=PatternFill("solid", fgColor="FDE2E2"),
                           font=Font(color=DANGER, bold=True)),
            )

        # ── Autofilter over the table (header + body) ────────────────────
        ws.auto_filter.ref = f"A{head_row}:{get_column_letter(ncols)}{last_data}"

    # ── Footer note ──────────────────────────────────────────────────────
    note_row = (last_data + 3) if rows else (head_row + 2)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=ncols)
    nc = ws.cell(row=note_row, column=1,
                 value="Confidential · Statutory filing document · Fourreck Technologies Pvt. Ltd. · "
                       "PF (EPFO) · ESI (ESIC) · Professional Tax · TDS (Form 24Q)")
    nc.font = Font(italic=True, size=7.5, color=BRONZE_DEEP)
    nc.alignment = Alignment(horizontal="left", vertical="center")

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()
